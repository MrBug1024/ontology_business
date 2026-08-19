# -*- coding: utf-8 -*-
import os
import openpyxl

base = r'E:\gx\new_docs\逻辑资料V1\AI智能体相关法律、准则、底稿、案例、报告模版【20260703】\7、案例'
case = [d for d in os.listdir(base) if '3999' in d][0]
mat = os.path.join(base, case)
for d in os.listdir(mat):
    if '拆分' in d:
        mat = os.path.join(mat, d)

files = ['基础信息.xlsx', '科目余额表-试算平衡派生.xlsx', '调整明细-规范化.xlsx',
         'C5-2应收帐款审定表.xlsx', 'F1-2主营业务收入审定表.xlsx',
         '银行流水-底稿抽查派生.xlsx', '工资.xlsx', '预收替代.xlsx',
         '营业外收入明细表-规范化.xlsx', '营业外支出明细表-规范化.xlsx',
         '咨询费合同抽查表.xlsx', 'F1-3主营业务收入检查表.xlsx']

for f in files:
    p = os.path.join(mat, f)
    if not os.path.exists(p):
        print('MISSING', f)
        continue
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    print('=' * 90)
    print(f, '| sheets:', wb.sheetnames)
    for ws in wb.worksheets[:2]:
        print('  -- sheet:', ws.title)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 10:
                break
            vals = [str(v)[:16] if v is not None else '' for v in row[:14]]
            print('   ', ' | '.join(vals))
    wb.close()
