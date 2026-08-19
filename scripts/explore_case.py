# -*- coding: utf-8 -*-
"""探查年审案例资料：底稿 xlsx 的 sheet 清单 + 关键 sheet 内容"""
import openpyxl, xlrd, os, sys

BASE = r"E:\gx\new_docs\逻辑资料V1\AI智能体相关法律、准则、底稿、案例、报告模版【20260703】\7、案例\北京有限公司-京创会审字[2024]第3999号-标准无保留意见-4.3.V2"

def dump_xlsx_sheets(path, max_rows=8, max_cols=12):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"\n########## {os.path.basename(path)}  sheets={len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  - {name}  ({ws.max_row}x{ws.max_column})")
    wb.close()

def dump_sheet(path, sheet, max_rows=30, max_cols=14):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    print(f"\n===== SHEET: {sheet} =====")
    for i, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True)):
        vals = ["" if v is None else str(v)[:20] for v in row]
        if any(v.strip() for v in vals):
            print(f"  r{i+1}: " + " | ".join(vals))
    wb.close()

def dump_xls(path, max_rows=25, max_cols=14):
    book = xlrd.open_workbook(path)
    print(f"\n########## {os.path.basename(path)}  sheets={book.nsheets}")
    for sh in book.sheets():
        print(f"  - {sh.name}  ({sh.nrows}x{sh.ncols})")
        for r in range(min(sh.nrows, max_rows)):
            vals = [str(sh.cell_value(r, c))[:18] for c in range(min(sh.ncols, max_cols))]
            if any(v.strip() for v in vals):
                print(f"    r{r+1}: " + " | ".join(vals))

if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "list"
    if mode == "list":
        dump_xlsx_sheets(os.path.join(BASE, "北京有限公司2023年年审底稿.xlsx"))
    elif mode == "sheet":
        dump_sheet(os.path.join(BASE, "北京有限公司2023年年审底稿.xlsx"), sys.argv[2],
                   int(sys.argv[3]) if len(sys.argv) > 3 else 30)
    elif mode == "xls":
        dump_xls(sys.argv[2])
