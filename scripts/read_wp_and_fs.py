# -*- coding: utf-8 -*-
import os
import openpyxl

root = r'E:\gx\new_docs\逻辑资料V1\AI智能体相关法律、准则、底稿、案例、报告模版【20260703】'

# 1. 审计底稿文件夹结构
print('=== 6、年度财务报表审计底稿 结构 ===')
wp = os.path.join(root, '6、年度财务报表审计底稿')
for r, d, fs in os.walk(wp):
    depth = r.replace(wp, '').count(os.sep)
    for f in fs:
        print('  ' * depth, f)

# 2. 经审计的财务报表.xls
print()
print('=== 经审计的财务报表.xls ===')
case = [x for x in os.listdir(os.path.join(root, '7、案例')) if '3999' in x][0]
mat = os.path.join(root, '7、案例', case)
for dd in os.listdir(mat):
    if '拆分' in dd:
        mat = os.path.join(mat, dd)
fs_path = os.path.join(mat, '财务报表', '2、经审计的财务报表-（会计准则）.xls')
try:
    wb = openpyxl.load_workbook(fs_path, read_only=True, data_only=True)
    print('sheets:', wb.sheetnames)
    for ws in wb.worksheets:
        print(f'-- {ws.title}')
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 12:
                break
            vals = [str(v)[:14] if v is not None else '' for v in row[:8]]
            print('   ', ' | '.join(vals))
    wb.close()
except Exception as e:
    print('xls err (old format):', e)
