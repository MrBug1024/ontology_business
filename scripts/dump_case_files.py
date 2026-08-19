# -*- coding: utf-8 -*-
"""dump 年审案例关键文件内容，用于完善代理记账业务场景。"""
import os, sys, json

BASE = r"E:\gx\new_docs\逻辑资料V1\AI智能体相关法律、准则、底稿、案例、报告模版【20260703】\7、案例\北京有限公司-京创会审字[2024]第3999号-标准无保留意见-4.3.V2"
SPLIT = os.path.join(BASE, "拆分上传材料")

def dump_xlsx(path, max_rows=60, max_cols=12):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"=== SHEET: {ws.title} ({ws.max_row}x{ws.max_column}) ===")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                out.append("... (truncated)")
                break
            vals = [("" if v is None else str(v))[:40] for v in row[:max_cols]]
            if any(v.strip() for v in vals):
                out.append(" | ".join(vals))
    return "\n".join(out)

def dump_xls(path, max_rows=60, max_cols=12):
    import xlrd
    wb = xlrd.open_workbook(path)
    out = []
    for ws in wb.sheets():
        out.append(f"=== SHEET: {ws.name} ({ws.nrows}x{ws.ncols}) ===")
        for i in range(min(ws.nrows, max_rows)):
            vals = [str(ws.cell_value(i, j))[:40] for j in range(min(ws.ncols, max_cols))]
            if any(v.strip() for v in vals):
                out.append(" | ".join(vals))
    return "\n".join(out)

def dump_doc(path, max_chars=8000):
    # .doc 老格式，尝试用 textract 不可用；用 olefile 提取 WordDocument 流文本
    try:
        import olefile
        ole = olefile.OleFileIO(path)
        data = ole.openstream("WordDocument").read()
        # 简单提取：doc 文本多为 UTF-16LE 或 GBK
        txt = None
        for enc in ("utf-16-le", "gbk"):
            try:
                t = data.decode(enc, errors="ignore")
                # 过滤控制字符
                t = "".join(c for c in t if c.isprintable() or c in "\n\r\t")
                if len(t) > 200:
                    txt = t
                    break
            except Exception:
                pass
        return (txt or "(no text)")[:max_chars]
    except Exception as e:
        return f"(doc parse error: {e})"

def dump_docx(path, max_chars=8000):
    import docx
    d = docx.Document(path)
    txt = "\n".join(p.text for p in d.paragraphs if p.text.strip())
    return txt[:max_chars]

targets = sys.argv[1:] or [
    "基础信息.xlsx",
    "科目余额表-试算平衡派生.xlsx",
    "C5-2应收帐款审定表.xlsx",
    "F1-2主营业务收入审定表.xlsx",
    "F1-3主营业务收入检查表.xlsx",
    "调整明细-规范化.xlsx",
    "银行流水-底稿抽查派生.xlsx",
    "工资.xlsx",
    "营业外收入明细表-规范化.xlsx",
    "营业外支出明细表-规范化.xlsx",
    "预收替代.xlsx",
    "咨询费合同抽查表.xlsx",
]
for t in targets:
    p = os.path.join(SPLIT, t)
    print("\n" + "=" * 80)
    print("FILE:", t)
    print("=" * 80)
    if not os.path.exists(p):
        print("(missing)")
        continue
    try:
        if t.endswith(".xlsx"):
            print(dump_xlsx(p))
        elif t.endswith(".xls"):
            print(dump_xls(p))
        elif t.endswith(".docx"):
            print(dump_docx(p))
        elif t.endswith(".doc"):
            print(dump_doc(p))
    except Exception as e:
        print(f"(error: {e})")
