from __future__ import annotations

from contextlib import nullcontext
from datetime import datetime, timedelta, timezone
from io import BytesIO
import hashlib
from pathlib import Path
import re
from types import SimpleNamespace
import unittest
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

import pyarrow.parquet as parquet
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.catalog_schemas import ValidationDatasetBuildIn
from app.database import Base
from app.models import (
    BucketFile,
    DataAsset,
    DataAssetVersion,
    DataSource,
    DatasetFragment,
    DatasetVersion,
    DocumentChunk,
    IngestionRun,
    LogicalDataset,
    Tenant,
    User,
)
from app.services import (
    catalog_ingestion_service,
    dataset_query_service,
    datasource_service,
    object_deletion_service,
    object_storage_service,
    permission_service,
    validation_dataset_service,
)


class ValidationDatasetServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        self.Session = sessionmaker(
            bind=self.engine,
            autoflush=False,
            expire_on_commit=False,
        )
        Base.metadata.create_all(self.engine)
        self.raw_objects: dict[tuple[str, str], bytes] = {}
        self.parquet_objects: dict[tuple[str, str], bytes] = {}
        content = (
            "claim_id,amount,approved\n"
            "12345678901234567890,12.50,true\n"
            "00000123,7,false\n"
        ).encode()
        digest = hashlib.sha256(content).hexdigest()
        with self.Session() as db:
            tenant = Tenant(id="tenant-validation-data", name="Validation data")
            user = User(
                id="user-validation-data",
                tenant_id=tenant.id,
                email="validation-data@example.test",
                password_hash="test-only",
                status="active",
            )
            source = DataSource(
                id="validation-managed-bucket",
                tenant_id=tenant.id,
                name="Validation managed bucket",
                type="file_bucket",
                config={
                    "storage_backend": "minio",
                    "bucket_name": "validation-test",
                    "prefix": "platform",
                },
            )
            bucket_file = BucketFile(
                id="raw-validation-csv",
                data_source_id=source.id,
                filename="claims.csv",
                stored_path="minio://validation-test/platform/raw/claims.csv",
                storage_provider="minio",
                bucket_name="validation-test",
                object_key="platform/raw/claims.csv",
                object_version_id="raw-v1",
                object_url="minio://validation-test/platform/raw/claims.csv",
                size=len(content),
                mime="text/csv",
                content_sha256=digest,
                status="parsed",
            )
            asset = DataAsset(
                id="validation-csv-asset",
                tenant_id=tenant.id,
                key="validation.claims",
                name="claims.csv",
                media_type="text/csv",
                labels={"catalog_purpose": "validation_asset"},
                created_by_user_id=user.id,
            )
            version = DataAssetVersion(
                id="validation-csv-version",
                tenant_id=tenant.id,
                asset_id=asset.id,
                version_number=1,
                bucket_file_id=bucket_file.id,
                bucket_data_source_id=source.id,
                status="ready",
                content_sha256=digest,
                byte_size=len(content),
                version_document={
                    "profile": {
                        "category": "table",
                        "tables": [
                            {
                                "name": "claims",
                                "header_row_index": 0,
                                "columns": [
                                    {"name": "claim_id", "logical_type": "integer"},
                                    {"name": "amount", "logical_type": "number"},
                                    {"name": "approved", "logical_type": "boolean"},
                                ],
                            }
                        ],
                    }
                },
                created_by_user_id=user.id,
            )
            db.add_all([tenant, user, source, bucket_file, asset, version])
            db.commit()
            permission_service.ensure_organization(
                db,
                tenant.id,
                owner_user_id=user.id,
            )
            db.commit()
        self.source_id = "validation-managed-bucket"
        self.asset_version_id = "validation-csv-version"
        self.raw_objects[("validation-test", "platform/raw/claims.csv")] = content

    def tearDown(self) -> None:
        self.engine.dispose()

    def _database(self):
        db = self.Session()
        db.info["tenant_id"] = "tenant-validation-data"
        db.info["user_id"] = "user-validation-data"
        return db

    def _download(
        self,
        bucket_name: str,
        object_key: str,
        destination: str | Path,
        **_kwargs,
    ):
        content = self.raw_objects.get((bucket_name, object_key))
        if content is None:
            content = self.parquet_objects[(bucket_name, object_key)]
        Path(destination).write_bytes(content)
        return SimpleNamespace(size=len(content))

    def _save_parquet(
        self,
        source: DataSource,
        filename: str,
        source_path: str | Path,
        *,
        mime: str,
        stable_file_id: str,
        upload_object_key: str,
        content_sha256: str,
    ) -> BucketFile:
        content = Path(source_path).read_bytes()
        self.parquet_objects[("validation-test", upload_object_key)] = content
        return BucketFile(
            id=stable_file_id,
            data_source_id=source.id,
            filename=filename,
            stored_path=f"minio://validation-test/{upload_object_key}",
            storage_provider="minio",
            bucket_name="validation-test",
            object_key=upload_object_key,
            object_version_id="parquet-v1",
            etag=content_sha256,
            object_url=f"minio://validation-test/{upload_object_key}",
            size=len(content),
            mime=mime,
            content_sha256=content_sha256,
        )

    def _patches(self):
        with self._database() as db:
            source = db.get(DataSource, self.source_id)
            db.expunge(source)

        def prepare_claim(data_source, file_id, filename):
            return SimpleNamespace(
                object_key=f"platform/files/{file_id}/generations/{'a' * 32}/{filename}"
            )

        return (
            patch.object(
                catalog_ingestion_service,
                "require_external_upload_bucket",
                return_value=source,
            ),
            patch.object(
                object_storage_service,
                "download_object_to_file",
                side_effect=self._download,
            ),
            patch.object(
                datasource_service,
                "save_bucket_file_path",
                side_effect=self._save_parquet,
            ),
            patch.object(
                object_deletion_service,
                "prepare_bucket_file_upload",
                side_effect=prepare_claim,
            ),
            patch.object(
                object_deletion_service,
                "heartbeat_upload_intent",
                side_effect=lambda _claim: nullcontext(SimpleNamespace()),
            ),
            patch.object(object_deletion_service, "begin_upload_put"),
            patch.object(object_deletion_service, "assert_upload_active"),
            patch.object(object_deletion_service, "retain_bucket_file_upload"),
        )

    def test_materializes_queryable_minio_parquet_without_postgresql_rows(self) -> None:
        from contextlib import ExitStack

        payload = ValidationDatasetBuildIn(
            asset_version_ids=[self.asset_version_id],
            name="Claims validation package",
        )
        with ExitStack() as stack:
            for active_patch in self._patches():
                stack.enter_context(active_patch)
            with self._database() as db:
                first = validation_dataset_service.build_validation_dataset(db, payload)
                second = validation_dataset_service.build_validation_dataset(db, payload)

        self.assertFalse(first["reused"])
        self.assertTrue(second["reused"])
        self.assertEqual(first["dataset_version_id"], second["dataset_version_id"])
        self.assertEqual(first["record_count"], 2)
        self.assertEqual(first["relation_names"], ["claims"])
        self.assertEqual(len(self.parquet_objects), 1)
        stored_table = parquet.read_table(BytesIO(next(iter(self.parquet_objects.values()))))
        self.assertEqual(
            stored_table.column("claim_id").to_pylist(),
            ["12345678901234567890", "00000123"],
        )

        with self._database() as db:
            self.assertEqual(db.scalar(select(func.count(DocumentChunk.id))), 0)
            self.assertEqual(db.scalar(select(func.count(LogicalDataset.id))), 1)
            self.assertEqual(db.scalar(select(func.count(DatasetVersion.id))), 1)
            self.assertEqual(db.scalar(select(func.count(DatasetFragment.id))), 1)
            query_source = DataSource(
                id="validation-query-source",
                tenant_id="tenant-validation-data",
                name="Validation query source",
                type="dataset",
                config={
                    "dataset_id": first["dataset_id"],
                    "dataset_version_id": first["dataset_version_id"],
                },
            )
            db.add(query_source)
            db.commit()

        with patch.object(dataset_query_service, "SessionLocal", self.Session):
            catalog = dataset_query_service._load_catalog(query_source)
        self.assertEqual(catalog.dataset_version_id, first["dataset_version_id"])
        self.assertEqual(catalog.relations[0].row_count, 2)

    def test_durable_job_materializes_and_exposes_reusable_result(self) -> None:
        from contextlib import ExitStack

        payload = ValidationDatasetBuildIn(
            asset_version_ids=[self.asset_version_id],
            name="Durable validation package",
        )
        with self._database() as db:
            queued = validation_dataset_service.enqueue_validation_dataset_job(
                db,
                payload,
            )
            duplicate = validation_dataset_service.enqueue_validation_dataset_job(
                db,
                payload,
            )
        self.assertEqual(queued["status"], "queued")
        self.assertEqual(duplicate["id"], queued["id"])

        with ExitStack() as stack:
            for active_patch in self._patches():
                stack.enter_context(active_patch)
            stack.enter_context(
                patch.object(validation_dataset_service, "SessionLocal", self.Session)
            )
            self.assertTrue(
                validation_dataset_service.process_validation_dataset_job(queued["id"])
            )

        with self._database() as db:
            completed = validation_dataset_service.get_validation_dataset_job(
                db,
                queued["id"],
            )
        self.assertEqual(completed["status"], "succeeded")
        self.assertEqual(completed["result"]["record_count"], 2)
        self.assertEqual(
            completed["result"]["source_asset_version_ids"],
            [self.asset_version_id],
        )

    def test_active_validation_job_lease_is_not_recovered_or_reclaimed(self) -> None:
        payload = ValidationDatasetBuildIn(
            asset_version_ids=[self.asset_version_id],
            name="Active lease validation package",
        )
        with self._database() as db:
            queued = validation_dataset_service.enqueue_validation_dataset_job(
                db, payload
            )
            run = db.get(IngestionRun, queued["id"])
            run.status = "running"
            run.lease_token = "active-validation-lease"
            run.lease_expires_at = datetime.now(timezone.utc) + timedelta(minutes=5)
            db.commit()

        with patch.object(validation_dataset_service, "SessionLocal", self.Session):
            self.assertEqual(
                validation_dataset_service.recover_validation_dataset_jobs(), 0
            )
            self.assertFalse(
                validation_dataset_service.process_validation_dataset_job(queued["id"])
            )

        with self._database() as db:
            run = db.get(IngestionRun, queued["id"])
            self.assertEqual(run.status, "running")
            self.assertEqual(run.lease_token, "active-validation-lease")

    def test_only_expired_validation_job_lease_is_requeued(self) -> None:
        payload = ValidationDatasetBuildIn(
            asset_version_ids=[self.asset_version_id],
            name="Expired lease validation package",
        )
        with self._database() as db:
            queued = validation_dataset_service.enqueue_validation_dataset_job(
                db, payload
            )
            run = db.get(IngestionRun, queued["id"])
            run.status = "running"
            run.lease_token = "expired-validation-lease"
            run.lease_expires_at = datetime.now(timezone.utc) - timedelta(seconds=1)
            db.commit()

        with patch.object(validation_dataset_service, "SessionLocal", self.Session):
            self.assertEqual(
                validation_dataset_service.recover_validation_dataset_jobs(), 1
            )

        with self._database() as db:
            run = db.get(IngestionRun, queued["id"])
            self.assertEqual(run.status, "pending")
            self.assertEqual(run.lease_token, "")
            self.assertIsNone(run.lease_expires_at)

    def test_releases_catalog_transaction_before_download_and_materialization(self) -> None:
        from contextlib import ExitStack

        payload = ValidationDatasetBuildIn(
            asset_version_ids=[self.asset_version_id],
            name="Transaction boundary validation package",
        )
        transaction_states: list[bool] = []
        with ExitStack() as stack:
            for active_patch in self._patches():
                stack.enter_context(active_patch)
            with self._database() as db:
                def download_without_open_catalog_transaction(*args, **kwargs):
                    transaction_states.append(db.in_transaction())
                    return self._download(*args, **kwargs)

                stack.enter_context(
                    patch.object(
                        object_storage_service,
                        "download_object_to_file",
                        side_effect=download_without_open_catalog_transaction,
                    )
                )
                validation_dataset_service.build_validation_dataset(db, payload)

        self.assertEqual(transaction_states, [False])

    def test_streaming_xlsx_without_dimension_materializes_all_profiled_sheets(self) -> None:
        from contextlib import ExitStack
        from openpyxl import Workbook

        workbook = Workbook()
        charge = workbook.active
        charge.title = "charge"
        charge.append(["charge_id", "amount"])
        charge.append(["C-1", 12.5])
        encounter = workbook.create_sheet("encounter")
        encounter.append(["encounter_id", "days"])
        encounter.append(["E-1", 3])
        original = BytesIO()
        workbook.save(original)
        workbook.close()

        rewritten = BytesIO()
        with ZipFile(BytesIO(original.getvalue())) as source, ZipFile(
            rewritten,
            "w",
            ZIP_DEFLATED,
        ) as destination:
            for member in source.infolist():
                content = source.read(member.filename)
                if member.filename.startswith("xl/worksheets/sheet"):
                    content = re.sub(rb"<dimension\b[^>]*/>", b"", content)
                destination.writestr(member, content)
        xlsx = rewritten.getvalue()
        digest = hashlib.sha256(xlsx).hexdigest()
        with self._database() as db:
            source = db.get(DataSource, self.source_id)
            bucket_file = BucketFile(
                id="raw-validation-xlsx",
                data_source_id=source.id,
                filename="validation.xlsx",
                stored_path="minio://validation-test/platform/raw/validation.xlsx",
                storage_provider="minio",
                bucket_name="validation-test",
                object_key="platform/raw/validation.xlsx",
                object_version_id="raw-xlsx-v1",
                object_url="minio://validation-test/platform/raw/validation.xlsx",
                size=len(xlsx),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                content_sha256=digest,
                status="parsed",
            )
            asset = DataAsset(
                id="validation-xlsx-asset",
                tenant_id="tenant-validation-data",
                key="validation.workbook",
                name="validation.xlsx",
                media_type=bucket_file.mime,
                labels={"catalog_purpose": "validation_asset"},
                created_by_user_id="user-validation-data",
            )
            version = DataAssetVersion(
                id="validation-xlsx-version",
                tenant_id="tenant-validation-data",
                asset_id=asset.id,
                version_number=1,
                bucket_file_id=bucket_file.id,
                bucket_data_source_id=source.id,
                status="ready",
                content_sha256=digest,
                byte_size=len(xlsx),
                version_document={
                    "profile": {
                        "category": "table",
                        "tables": [
                            {
                                "name": "charge",
                                "relation_name": "validation__charge",
                                "header_row_index": 0,
                                "columns": [
                                    {"name": "charge_id", "logical_type": "string"},
                                    {"name": "amount", "logical_type": "number"},
                                ],
                            },
                            {
                                "name": "encounter",
                                "relation_name": "validation__encounter",
                                "header_row_index": 0,
                                "columns": [
                                    {"name": "encounter_id", "logical_type": "string"},
                                    {"name": "days", "logical_type": "integer"},
                                ],
                            },
                        ],
                    }
                },
                created_by_user_id="user-validation-data",
            )
            db.add_all([bucket_file, asset, version])
            db.commit()
        self.raw_objects[("validation-test", "platform/raw/validation.xlsx")] = xlsx

        payload = ValidationDatasetBuildIn(
            asset_version_ids=[version.id],
            name="Dimensionless workbook validation package",
        )
        with ExitStack() as stack:
            for active_patch in self._patches():
                stack.enter_context(active_patch)
            with self._database() as db:
                result = validation_dataset_service.build_validation_dataset(db, payload)

        self.assertEqual(
            result["relation_names"],
            ["validation__charge", "validation__encounter"],
        )
        self.assertEqual(result["record_count"], 2)

    def test_package_keeps_queryable_tables_when_an_output_template_has_no_rows(self) -> None:
        from contextlib import ExitStack
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "result"
        sheet.append(["finding_id", "reason"])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        template = output.getvalue()
        digest = hashlib.sha256(template).hexdigest()
        with self._database() as db:
            source = db.get(DataSource, self.source_id)
            bucket_file = BucketFile(
                id="raw-validation-output-template",
                data_source_id=source.id,
                filename="result-template.xlsx",
                stored_path="minio://validation-test/platform/raw/result-template.xlsx",
                storage_provider="minio",
                bucket_name="validation-test",
                object_key="platform/raw/result-template.xlsx",
                object_version_id="raw-output-template-v1",
                object_url="minio://validation-test/platform/raw/result-template.xlsx",
                size=len(template),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                content_sha256=digest,
                status="parsed",
            )
            asset = DataAsset(
                id="validation-output-template-asset",
                tenant_id="tenant-validation-data",
                key="validation.output-template",
                name="result-template.xlsx",
                media_type=bucket_file.mime,
                labels={"catalog_purpose": "validation_asset"},
                created_by_user_id="user-validation-data",
            )
            version = DataAssetVersion(
                id="validation-output-template-version",
                tenant_id="tenant-validation-data",
                asset_id=asset.id,
                version_number=1,
                bucket_file_id=bucket_file.id,
                bucket_data_source_id=source.id,
                status="ready",
                content_sha256=digest,
                byte_size=len(template),
                version_document={
                    "profile": {
                        "category": "table",
                        "tables": [
                            {
                                "name": "result",
                                "relation_name": "result_template",
                                "header_row_index": 0,
                                "columns": [
                                    {"name": "finding_id", "logical_type": "string"},
                                    {"name": "reason", "logical_type": "string"},
                                ],
                            }
                        ],
                    }
                },
                created_by_user_id="user-validation-data",
            )
            db.add_all([bucket_file, asset, version])
            db.commit()
        self.raw_objects[
            ("validation-test", "platform/raw/result-template.xlsx")
        ] = template

        payload = ValidationDatasetBuildIn(
            asset_version_ids=[self.asset_version_id, version.id],
            name="Validation package with output template",
        )
        with ExitStack() as stack:
            for active_patch in self._patches():
                stack.enter_context(active_patch)
            with self._database() as db:
                result = validation_dataset_service.build_validation_dataset(db, payload)
            with self._database() as db:
                with self.assertRaisesRegex(
                    validation_dataset_service.ValidationDatasetError,
                    "没有可物化的表格数据",
                ):
                    validation_dataset_service.build_validation_dataset(
                        db,
                        ValidationDatasetBuildIn(
                            asset_version_ids=[version.id],
                            name="Output template only",
                        ),
                    )

        self.assertEqual(result["relation_names"], ["claims"])
        self.assertEqual(result["record_count"], 2)


if __name__ == "__main__":
    unittest.main()
