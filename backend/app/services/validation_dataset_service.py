"""Materialize reusable validation tables as MinIO-backed Parquet datasets."""
from __future__ import annotations

import csv
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
import hashlib
from itertools import islice
import json
import math
from pathlib import Path
import re
import tempfile
from typing import Any, Iterable, Iterator, Sequence
import uuid

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..catalog_schemas import ValidationDatasetBuildIn
from ..config import get_settings
from ..database import SessionLocal
from ..models import (
    BucketFile,
    DataAsset,
    DataAssetVersion,
    DatasetField,
    DatasetFragment,
    DatasetRelation,
    DatasetSchema,
    DatasetVersion,
    DatasetVersionAsset,
    LogicalDataset,
)
from . import (
    catalog_ingestion_service,
    datasource_service,
    object_deletion_service,
    object_storage_service,
    permission_service,
    tenant_service,
)


class ValidationDatasetError(ValueError):
    pass


_TABLE_EXTENSIONS = {".csv", ".tsv", ".xls", ".xlsx", ".xlsm"}
_IDENTIFIER_MARKERS = (
    "id", "code", "number", "编号", "编码", "代码", "证件", "卡号", "单号", "序号"
)
_KEY_CLEAN_RE = re.compile(r"[\x00-\x1f\x7f]")
_BATCH_ROWS = 5_000
_PARQUET_TARGET_BYTES = 256 * 1024 * 1024


def _canonical_hash(value: Any) -> str:
    encoded = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
        default=str,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _file_sha256(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(4 * 1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _relation_key(value: str, fallback: str) -> str:
    key = _KEY_CLEAN_RE.sub("", str(value or "")).strip()
    if not key:
        key = fallback
    if len(key) > 180:
        key = f"{key[:140]}-{hashlib.sha256(key.encode()).hexdigest()[:16]}"
    return key


def _unique_name(value: Any, ordinal: int, seen: set[str]) -> str:
    base = str(value).strip() if value is not None else ""
    base = _KEY_CLEAN_RE.sub("", base)[:300] or f"column_{ordinal + 1}"
    candidate = base
    suffix = 2
    while candidate.casefold() in seen:
        candidate = f"{base}_{suffix}"
        suffix += 1
    seen.add(candidate.casefold())
    return candidate


def _logical_type(profile_type: str, name: str) -> str:
    normalized = str(profile_type or "string").strip().lower()
    folded = name.casefold()
    if normalized in {"integer", "number"} and any(
        marker in folded for marker in _IDENTIFIER_MARKERS
    ):
        return "string"
    return normalized if normalized in {
        "integer", "number", "boolean", "date", "datetime", "string"
    } else "string"


def _arrow_contract(columns: Sequence[dict[str, Any]]):
    import pyarrow as pa

    arrow_fields = []
    field_contract = []
    logical_types: list[str] = []
    physical = {
        "integer": (pa.int64(), "BIGINT"),
        "number": (pa.float64(), "DOUBLE"),
        "boolean": (pa.bool_(), "BOOLEAN"),
        "date": (pa.date32(), "DATE"),
        "datetime": (pa.timestamp("us"), "TIMESTAMP"),
        "string": (pa.string(), "VARCHAR"),
    }
    seen: set[str] = set()
    for ordinal, column in enumerate(columns):
        name = _unique_name(column.get("name"), ordinal, seen)
        logical = _logical_type(str(column.get("logical_type") or ""), name)
        arrow_type, sql_type = physical[logical]
        arrow_fields.append(pa.field(name, arrow_type, nullable=True))
        logical_types.append(logical)
        field_contract.append(
            {
                "name": name,
                "physical_type": sql_type,
                "nullable": True,
                "key_ordinal": None,
                "ordinal": ordinal,
                "logical_type": logical,
            }
        )
    if not arrow_fields:
        raise ValidationDatasetError("表格没有可物化的字段")
    return pa.schema(arrow_fields), field_contract, logical_types


def _coerce(value: Any, logical: str) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if logical == "string":
        return str(value)
    if logical == "integer":
        try:
            parsed = int(Decimal(str(value).strip()))
            return parsed if -(2**63) <= parsed < 2**63 else None
        except (InvalidOperation, ValueError, TypeError, OverflowError):
            return None
    if logical == "number":
        try:
            parsed = float(Decimal(str(value).strip()))
            return parsed if math.isfinite(parsed) else None
        except (InvalidOperation, ValueError, TypeError, OverflowError):
            return None
    if logical == "boolean":
        if isinstance(value, bool):
            return value
        lowered = str(value).strip().casefold()
        if lowered in {"true", "1", "yes", "y", "是"}:
            return True
        if lowered in {"false", "0", "no", "n", "否"}:
            return False
        return None
    if logical == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return date.fromisoformat(str(value).strip()[:10])
        except ValueError:
            return None
    if logical == "datetime":
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        try:
            return datetime.fromisoformat(str(value).strip().replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            return None
    return str(value)


def _write_parquet_fragments(
    output_dir: Path,
    output_stem: str,
    rows: Iterable[Sequence[Any]],
    schema: Any,
    logical_types: Sequence[str],
) -> list[dict[str, Any]]:
    import pyarrow as pa
    import pyarrow.parquet as parquet

    writer = None
    destination: Path | None = None
    fragment_rows = 0
    fragment_index = 0
    fragments: list[dict[str, Any]] = []
    batch: list[Sequence[Any]] = []

    def close_fragment() -> None:
        nonlocal writer, destination, fragment_rows
        if writer is None or destination is None:
            return
        writer.close()
        digest = _file_sha256(destination)
        fragments.append(
            {
                "path": destination,
                "row_count": fragment_rows,
                "byte_size": destination.stat().st_size,
                "content_sha256": digest,
            }
        )
        writer = None
        destination = None
        fragment_rows = 0

    def flush() -> None:
        nonlocal writer, destination, fragment_rows, fragment_index
        if not batch:
            return
        if writer is None:
            fragment_index += 1
            destination = output_dir / f"{output_stem}-{fragment_index:04d}.parquet"
            writer = parquet.ParquetWriter(destination, schema, compression="zstd")
        arrays = []
        for index, field in enumerate(schema):
            arrays.append(
                pa.array(
                    [
                        _coerce(row[index] if index < len(row) else None, logical_types[index])
                        for row in batch
                    ],
                    type=field.type,
                )
            )
        writer.write_batch(pa.RecordBatch.from_arrays(arrays, schema=schema))
        fragment_rows += len(batch)
        batch.clear()
        if destination.stat().st_size >= _PARQUET_TARGET_BYTES:
            close_fragment()

    try:
        for row in rows:
            values = tuple(row)
            if not any(value is not None and str(value).strip() for value in values):
                continue
            batch.append(values)
            if len(batch) >= _BATCH_ROWS:
                flush()
        flush()
    finally:
        close_fragment()
    return fragments


def _profile(version: DataAssetVersion) -> dict[str, Any]:
    document = version.version_document or {}
    profile = document.get("profile") if isinstance(document, dict) else None
    if not isinstance(profile, dict) or profile.get("category") != "table":
        raise ValidationDatasetError("所选资料不是可查询表格")
    return profile


def _csv_rows(path: Path, delimiter: str) -> Iterator[Sequence[Any]]:
    with path.open("rb") as handle:
        sample = handle.read(65536)
    encoding = "utf-8-sig"
    for candidate in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            sample.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    with path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        next(reader, None)
        yield from reader


def _close_tabular_workbook(workbook: Any) -> None:
    if workbook is None:
        return
    close = getattr(workbook, "close", None)
    release = getattr(workbook, "release_resources", None)
    if callable(close):
        close()
    elif callable(release):
        release()


def _materialize_raw_file(
    raw_path: Path,
    filename: str,
    profile: dict[str, Any],
    output_dir: Path,
    used_relation_keys: set[str],
) -> list[dict[str, Any]]:
    extension = Path(filename).suffix.lower()
    stem = Path(filename).stem
    tables = list(profile.get("tables") or [])
    results: list[dict[str, Any]] = []
    if extension in {".csv", ".tsv"}:
        if not tables:
            raise ValidationDatasetError("表格结构 profile 缺失")
        relation_name = str(tables[0].get("relation_name") or "").strip()
        candidates = [(
            relation_name
            or catalog_ingestion_service.runtime_relation_name(filename, "data", 1),
            tables[0],
            _csv_rows(raw_path, "\t" if extension == ".tsv" else ","),
        )]
        workbook = None
    elif extension in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(
            raw_path,
            read_only=True,
            data_only=True,
            keep_links=False,
            keep_vba=False,
        )
        profile_by_name = {str(item.get("name") or ""): item for item in tables}
        candidates = []
        try:
            # OOXML may omit worksheet dimensions. In read-only mode openpyxl
            # then exposes max_row=None, so the verified profile is authoritative.
            profiled_sheets = [
                sheet for sheet in workbook.worksheets
                if sheet.title in profile_by_name
            ]
            for sheet in profiled_sheets:
                table_profile = profile_by_name[sheet.title]
                rows = sheet.iter_rows(values_only=True)
                rows = islice(
                    rows,
                    int(table_profile.get("header_row_index") or 0) + 1,
                    None,
                )
                relation_name = str(table_profile.get("relation_name") or "").strip()
                if not relation_name:
                    relation_name = catalog_ingestion_service.runtime_relation_name(
                        filename, sheet.title, len(profiled_sheets)
                    )
                candidates.append((relation_name, table_profile, rows))
        except Exception:
            _close_tabular_workbook(workbook)
            raise
    elif extension == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(str(raw_path), on_demand=True)
        try:
            profile_by_name = {str(item.get("name") or ""): item for item in tables}
            candidates = []
            nonempty_sheets = [
                workbook.sheet_by_index(index)
                for index in range(workbook.nsheets)
                if workbook.sheet_by_index(index).nrows > 0
            ]
            for sheet in nonempty_sheets:
                table_profile = profile_by_name.get(sheet.name)
                if table_profile is None:
                    continue
                start = int(table_profile.get("header_row_index") or 0) + 1
                rows = (sheet.row_values(index) for index in range(start, sheet.nrows))
                relation_name = str(table_profile.get("relation_name") or "").strip()
                if not relation_name:
                    relation_name = catalog_ingestion_service.runtime_relation_name(
                        filename, sheet.name, len(nonempty_sheets)
                    )
                candidates.append((relation_name, table_profile, rows))
        except Exception:
            _close_tabular_workbook(workbook)
            raise
    else:
        raise ValidationDatasetError("验证数据包目前只接受 CSV、TSV、XLS 或 XLSX")
    try:
        for ordinal, (raw_relation_name, table_profile, rows) in enumerate(candidates):
            relation_name = _relation_key(raw_relation_name, f"table_{ordinal + 1}")
            if relation_name.casefold() in used_relation_keys:
                relation_name = _relation_key(
                    f"{stem}__{relation_name}", f"table_{len(used_relation_keys) + 1}"
                )
            if relation_name.casefold() in used_relation_keys:
                raise ValidationDatasetError(f"验证数据包包含重复关系名: {relation_name}")
            used_relation_keys.add(relation_name.casefold())
            arrow_schema, field_contract, logical_types = _arrow_contract(
                list(table_profile.get("columns") or [])
            )
            fragments = _write_parquet_fragments(
                output_dir,
                f"{len(used_relation_keys):04d}",
                rows,
                arrow_schema,
                logical_types,
            )
            if not fragments:
                continue
            results.append(
                {
                    "relation_key": relation_name,
                    "display_name": relation_name,
                    "fields": field_contract,
                    "fragments": fragments,
                    "row_count": sum(item["row_count"] for item in fragments),
                    "byte_size": sum(item["byte_size"] for item in fragments),
                }
            )
    finally:
        _close_tabular_workbook(workbook)
    if not results:
        raise ValidationDatasetError(f"{filename} 没有可物化的工作表")
    return results


def _result(dataset: LogicalDataset, version: DatasetVersion, schema: DatasetSchema, *, reused: bool) -> dict[str, Any]:
    relations = list(
        version.manifest.get("relations", {}).keys()
        if isinstance(version.manifest, dict)
        else []
    )
    return {
        "dataset_id": dataset.id,
        "dataset_version_id": version.id,
        "content_hash": version.content_hash,
        "schema_hash": schema.schema_hash,
        "record_count": version.record_count,
        "byte_size": version.byte_size,
        "relation_names": relations,
        "source_asset_version_ids": [],
        "reused": reused,
    }


def _input_identity(versions: Sequence[DataAssetVersion]) -> tuple[dict[str, Any], str, str]:
    identity = {
        "format": "validation-dataset-input/v1",
        "assets": [
            {"id": item.id, "sha256": item.content_sha256}
            for item in versions
        ],
    }
    identity_hash = _canonical_hash(identity)
    return identity, identity_hash, f"validation.bundle.{identity_hash[:32]}"


def _job_document(db: Session, dataset: LogicalDataset) -> dict[str, Any]:
    labels = dict(dataset.labels or {})
    version = db.scalar(
        select(DatasetVersion)
        .where(
            DatasetVersion.dataset_id == dataset.id,
            DatasetVersion.status == "ready",
        )
        .order_by(DatasetVersion.version_number.desc())
    )
    status = "succeeded" if version is not None else str(
        labels.get("build_status") or "queued"
    )
    result = None
    if version is not None:
        schema = db.get(DatasetSchema, version.schema_id)
        if schema is not None:
            result = _result(dataset, version, schema, reused=True)
            result["source_asset_version_ids"] = list(
                labels.get("source_asset_version_ids") or []
            )
    return {
        "id": dataset.id,
        "status": status,
        "error": str(labels.get("build_error") or ""),
        "created_at": dataset.created_at,
        "updated_at": dataset.updated_at,
        "result": result,
    }


def enqueue_validation_dataset_job(
    db: Session,
    payload: ValidationDatasetBuildIn,
) -> dict[str, Any]:
    permission_service.require_tenant_permission(db, "write")
    tenant_id = tenant_service.current_tenant_id(db)
    versions = list(
        db.scalars(
            select(DataAssetVersion).where(
                DataAssetVersion.id.in_(payload.asset_version_ids),
                DataAssetVersion.tenant_id == tenant_id,
                DataAssetVersion.status == "ready",
                DataAssetVersion.bucket_file_id.is_not(None),
            )
        ).all()
    )
    by_id = {item.id: item for item in versions}
    if len(by_id) != len(payload.asset_version_ids):
        raise ValidationDatasetError("部分验证资料不存在、已删除或尚未就绪")
    ordered = [by_id[item] for item in payload.asset_version_ids]
    _identity, identity_hash, dataset_key = _input_identity(ordered)
    dataset = db.scalar(
        select(LogicalDataset)
        .where(
            LogicalDataset.tenant_id == tenant_id,
            LogicalDataset.key == dataset_key,
        )
        .with_for_update()
    )
    if dataset is None:
        dataset = LogicalDataset(
            tenant_id=tenant_id,
            key=dataset_key,
            name=payload.name.strip(),
            description="验证中心按内容哈希生成的可复用数据包",
            lifecycle_status="active",
            labels={
                "catalog_purpose": "validation_dataset",
                "input_hash": identity_hash,
                "source_asset_version_ids": list(payload.asset_version_ids),
                "build_status": "queued",
                "build_error": "",
                "requested_by_user_id": str(db.info.get("user_id") or ""),
            },
            created_by_user_id=str(db.info.get("user_id") or "") or None,
        )
        db.add(dataset)
        db.flush()
    else:
        document = _job_document(db, dataset)
        if document["status"] == "failed":
            dataset.labels = {
                **dict(dataset.labels or {}),
                "build_status": "queued",
                "build_error": "",
            }
    db.commit()
    db.refresh(dataset)
    return _job_document(db, dataset)


def get_validation_dataset_job(db: Session, job_id: str) -> dict[str, Any]:
    permission_service.require_tenant_permission(db, "read")
    dataset = db.scalar(
        select(LogicalDataset).where(
            LogicalDataset.id == job_id,
            LogicalDataset.tenant_id == tenant_service.current_tenant_id(db),
        )
    )
    if dataset is None or (dataset.labels or {}).get("catalog_purpose") != "validation_dataset":
        raise ValidationDatasetError("验证数据集任务不存在")
    return _job_document(db, dataset)


def process_validation_dataset_job(job_id: str) -> bool:
    """Claim and execute one durable dataset job in a worker-owned session."""
    db = SessionLocal()
    try:
        dataset = db.scalar(
            select(LogicalDataset)
            .where(LogicalDataset.id == job_id)
            .with_for_update()
        )
        if dataset is None:
            return False
        labels = dict(dataset.labels or {})
        if labels.get("catalog_purpose") != "validation_dataset":
            return False
        if str(labels.get("build_status") or "") != "queued":
            return False
        tenant_id = str(dataset.tenant_id)
        user_id = str(labels.get("requested_by_user_id") or dataset.created_by_user_id or "")
        dataset.labels = {
            **labels,
            "build_status": "running",
            "build_started_at": datetime.now(timezone.utc).isoformat(),
            "build_error": "",
        }
        db.commit()
        db.info["tenant_id"] = tenant_id
        db.info["user_id"] = user_id
        payload = ValidationDatasetBuildIn(
            asset_version_ids=list(labels.get("source_asset_version_ids") or []),
            name=dataset.name,
        )
        build_validation_dataset(db, payload)
        completed = db.get(LogicalDataset, job_id)
        if completed is not None:
            completed.labels = {
                **dict(completed.labels or {}),
                "build_status": "succeeded",
                "build_completed_at": datetime.now(timezone.utc).isoformat(),
                "build_error": "",
            }
            db.commit()
        return True
    except Exception as exc:  # noqa: BLE001 - worker exposes only a stable error.
        db.rollback()
        failed = db.get(LogicalDataset, job_id)
        if failed is not None:
            failed.labels = {
                **dict(failed.labels or {}),
                "build_status": "failed",
                "build_completed_at": datetime.now(timezone.utc).isoformat(),
                "build_error": str(exc or "验证数据集生成失败")[:1000],
            }
            db.commit()
        return False
    finally:
        db.close()


def process_next_validation_dataset_job() -> bool:
    db = SessionLocal()
    try:
        candidates = list(
            db.scalars(
                select(LogicalDataset)
                .where(
                    LogicalDataset.lifecycle_status == "active",
                    LogicalDataset.labels["catalog_purpose"].as_string()
                    == "validation_dataset",
                    LogicalDataset.labels["build_status"].as_string() == "queued",
                )
                .order_by(LogicalDataset.created_at, LogicalDataset.id)
                .limit(1)
            ).all()
        )
        job_id = next(
            (
                item.id
                for item in candidates
                if (item.labels or {}).get("catalog_purpose") == "validation_dataset"
                and (item.labels or {}).get("build_status") == "queued"
            ),
            None,
        )
    finally:
        db.close()
    return process_validation_dataset_job(job_id) if job_id else False


def recover_validation_dataset_jobs() -> int:
    db = SessionLocal()
    recovered = 0
    try:
        for dataset in db.scalars(
            select(LogicalDataset).where(
                LogicalDataset.lifecycle_status == "active",
                LogicalDataset.labels["catalog_purpose"].as_string()
                == "validation_dataset",
                LogicalDataset.labels["build_status"].as_string() == "running",
            )
        ):
            labels = dict(dataset.labels or {})
            if (
                labels.get("catalog_purpose") == "validation_dataset"
                and labels.get("build_status") == "running"
            ):
                dataset.labels = {
                    **labels,
                    "build_status": "queued",
                    "build_error": "服务重启后自动恢复",
                }
                recovered += 1
        db.commit()
        return recovered
    finally:
        db.close()


def build_validation_dataset(
    db: Session,
    payload: ValidationDatasetBuildIn,
) -> dict[str, Any]:
    permission_service.require_tenant_permission(db, "write")
    tenant_id = tenant_service.current_tenant_id(db)
    versions = list(
        db.scalars(
            select(DataAssetVersion)
            .where(
                DataAssetVersion.id.in_(payload.asset_version_ids),
                DataAssetVersion.tenant_id == tenant_id,
                DataAssetVersion.status == "ready",
                DataAssetVersion.bucket_file_id.is_not(None),
            )
        ).all()
    )
    by_id = {item.id: item for item in versions}
    if len(by_id) != len(payload.asset_version_ids):
        raise ValidationDatasetError("部分验证资料不存在、已删除或尚未就绪")
    versions = [by_id[item] for item in payload.asset_version_ids]
    assets = {
        item.id: item
        for item in db.scalars(
            select(DataAsset).where(
                DataAsset.id.in_([version.asset_id for version in versions]),
                DataAsset.tenant_id == tenant_id,
                DataAsset.lifecycle_status == "active",
            )
        ).all()
    }
    if len(assets) != len({item.asset_id for item in versions}):
        raise ValidationDatasetError("部分验证资料已删除")
    for version in versions:
        asset = assets[version.asset_id]
        if Path(asset.name).suffix.lower() not in _TABLE_EXTENSIONS:
            raise ValidationDatasetError(f"{asset.name} 不是可查询表格")
        _profile(version)

    identity, identity_hash, dataset_key = _input_identity(versions)
    existing_dataset = db.scalar(
        select(LogicalDataset).where(
            LogicalDataset.tenant_id == tenant_id,
            LogicalDataset.key == dataset_key,
        )
    )
    if existing_dataset is not None:
        existing_version = db.scalar(
            select(DatasetVersion)
            .where(
                DatasetVersion.dataset_id == existing_dataset.id,
                DatasetVersion.status == "ready",
            )
            .order_by(DatasetVersion.version_number.desc())
        )
        if existing_version is not None:
            existing_schema = db.get(DatasetSchema, existing_version.schema_id)
            if existing_schema is not None:
                result = _result(existing_dataset, existing_version, existing_schema, reused=True)
                result["source_asset_version_ids"] = list(payload.asset_version_ids)
                return result

    source = catalog_ingestion_service.require_external_upload_bucket(db)
    max_bytes = int(get_settings().catalog_max_upload_bytes)
    upload_records: list[tuple[Any, BucketFile]] = []
    try:
        with tempfile.TemporaryDirectory(prefix="ontology-validation-") as raw_dir:
            work = Path(raw_dir).resolve()
            materialized: list[dict[str, Any]] = []
            used_relation_keys: set[str] = set()
            for index, version in enumerate(versions):
                bucket_file = db.get(BucketFile, version.bucket_file_id)
                if bucket_file is None:
                    raise ValidationDatasetError("验证资料的 MinIO 对象记录不存在")
                raw_path = work / f"raw-{index:03d}{Path(assets[version.asset_id].name).suffix.lower()}"
                object_storage_service.download_object_to_file(
                    bucket_file.bucket_name,
                    bucket_file.object_key,
                    raw_path,
                    version_id=bucket_file.object_version_id,
                    max_bytes=max_bytes,
                )
                if _file_sha256(raw_path) != version.content_sha256:
                    raise ValidationDatasetError("验证资料完整性校验失败")
                materialized.extend(
                    _materialize_raw_file(
                        raw_path,
                        assets[version.asset_id].name,
                        _profile(version),
                        work,
                        used_relation_keys,
                    )
                )

            schema_relations = {
                item["relation_key"]: [
                    {
                        key: field[key]
                        for key in ("name", "physical_type", "nullable", "key_ordinal", "ordinal")
                    }
                    for field in item["fields"]
                ]
                for item in materialized
            }
            schema_document = {"relations": schema_relations, "derived_relations": {}}
            schema_hash = _canonical_hash(schema_document)
            if existing_dataset is None:
                existing_dataset = LogicalDataset(
                    tenant_id=tenant_id,
                    key=dataset_key,
                    name=payload.name.strip(),
                    description="验证中心按内容哈希生成的可复用数据包",
                    lifecycle_status="active",
                    labels={"catalog_purpose": "validation_dataset", "input_hash": identity_hash},
                    created_by_user_id=str(db.info.get("user_id") or "") or None,
                )
                db.add(existing_dataset)
                db.flush()
            schema = DatasetSchema(
                tenant_id=tenant_id,
                dataset_id=existing_dataset.id,
                schema_version=int(
                    db.scalar(
                        select(func.coalesce(func.max(DatasetSchema.schema_version), 0)).where(
                            DatasetSchema.dataset_id == existing_dataset.id
                        )
                    )
                    or 0
                ) + 1,
                schema_hash=schema_hash,
                compatibility="none",
                schema_document=schema_document,
                created_by_user_id=str(db.info.get("user_id") or "") or None,
            )
            db.add(schema)
            db.flush()
            relation_rows: list[tuple[DatasetRelation, dict[str, Any]]] = []
            for ordinal, item in enumerate(materialized):
                relation = DatasetRelation(
                    tenant_id=tenant_id,
                    dataset_id=existing_dataset.id,
                    schema_id=schema.id,
                    relation_key=item["relation_key"],
                    display_name=item["display_name"],
                    kind="table",
                    ordinal=ordinal,
                    description="",
                )
                db.add(relation)
                db.flush()
                for field in item["fields"]:
                    db.add(
                        DatasetField(
                            tenant_id=tenant_id,
                            dataset_id=existing_dataset.id,
                            schema_id=schema.id,
                            dataset_relation_id=relation.id,
                            field_key=f"field_{int(field['ordinal']) + 1}",
                            source_name=field["name"],
                            logical_type=field["logical_type"],
                            physical_type=field["physical_type"],
                            nullable=True,
                            ordinal=field["ordinal"],
                            key_ordinal=None,
                            semantic_role="",
                            field_document={},
                        )
                    )
                relation_rows.append((relation, item))
            db.flush()

            manifest_relations = {
                item["relation_key"]: {
                    "schema_hash": _canonical_hash(schema_relations[item["relation_key"]]),
                    "row_count": item["row_count"],
                    "byte_size": item["byte_size"],
                    **(
                        {"content_sha256": item["fragments"][0]["content_sha256"]}
                        if len(item["fragments"]) == 1
                        else {}
                    ),
                }
                for item in materialized
            }
            manifest = {
                "format": "validation-dataset/v1",
                "input": identity,
                "relations": manifest_relations,
                "derived_relations": {},
            }
            version_hash = _canonical_hash(
                {"dataset_key": dataset_key, "schema_hash": schema_hash, "manifest": manifest}
            )
            version = DatasetVersion(
                tenant_id=tenant_id,
                dataset_id=existing_dataset.id,
                schema_id=schema.id,
                version_number=int(
                    db.scalar(
                        select(func.coalesce(func.max(DatasetVersion.version_number), 0)).where(
                            DatasetVersion.dataset_id == existing_dataset.id
                        )
                    )
                    or 0
                ) + 1,
                status="ready",
                record_count=sum(int(item["row_count"]) for item in materialized),
                fragment_count=sum(len(item["fragments"]) for item in materialized),
                byte_size=sum(int(item["byte_size"]) for item in materialized),
                content_hash=version_hash,
                manifest=manifest,
                created_by_user_id=str(db.info.get("user_id") or "") or None,
                ready_at=datetime.now(timezone.utc),
            )
            db.add(version)
            db.flush()
            for ordinal, raw_version in enumerate(versions):
                db.add(
                    DatasetVersionAsset(
                        tenant_id=tenant_id,
                        dataset_id=existing_dataset.id,
                        dataset_version_id=version.id,
                        asset_version_id=raw_version.id,
                        role="source",
                        ordinal=ordinal,
                        binding_document={},
                    )
                )
            output_ordinal = 0
            for relation, item in relation_rows:
                for fragment_ordinal, fragment in enumerate(item["fragments"]):
                    output_ordinal += 1
                    file_id = uuid.uuid4().hex
                    filename = f"fragment-{output_ordinal:04d}.parquet"
                    claim = object_deletion_service.prepare_bucket_file_upload(
                        source, file_id, filename
                    )
                    with object_deletion_service.heartbeat_upload_intent(claim) as heartbeat:
                        object_deletion_service.begin_upload_put(claim)
                        parquet_file = datasource_service.save_bucket_file_path(
                            source,
                            filename,
                            fragment["path"],
                            mime="application/vnd.apache.parquet",
                            stable_file_id=file_id,
                            upload_object_key=claim.object_key,
                            content_sha256=fragment["content_sha256"],
                        )
                        object_deletion_service.assert_upload_active(
                            heartbeat, claim, parquet_file
                        )
                    parquet_file.status = "parsed"
                    parquet_file.index_status = "not_applicable"
                    db.add(parquet_file)
                    object_deletion_service.retain_bucket_file_upload(
                        db, claim, parquet_file, source
                    )
                    db.flush()
                    upload_records.append((claim, parquet_file))
                    db.add(
                        DatasetFragment(
                            tenant_id=tenant_id,
                            dataset_id=existing_dataset.id,
                            dataset_version_id=version.id,
                            dataset_relation_id=relation.id,
                            schema_id=schema.id,
                            bucket_file_id=parquet_file.id,
                            bucket_data_source_id=source.id,
                            ordinal=fragment_ordinal,
                            format="parquet",
                            compression="zstd",
                            status="ready",
                            row_count=fragment["row_count"],
                            byte_size=fragment["byte_size"],
                            content_sha256=fragment["content_sha256"],
                            statistics={"source_asset_count": len(versions)},
                        )
                    )
            db.commit()
            result = _result(existing_dataset, version, schema, reused=False)
            result["source_asset_version_ids"] = list(payload.asset_version_ids)
            return result
    except Exception:
        db.rollback()
        for claim, bucket_file in upload_records:
            object_deletion_service.schedule_abandoned_upload_best_effort(
                claim, bucket_file
            )
        raise


__all__ = [
    "ValidationDatasetError",
    "build_validation_dataset",
    "enqueue_validation_dataset_job",
    "get_validation_dataset_job",
    "process_next_validation_dataset_job",
    "process_validation_dataset_job",
    "recover_validation_dataset_jobs",
]
