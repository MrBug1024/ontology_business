"""Safe, generic profiling and persistence helpers for catalog file uploads.

The module deliberately understands file formats, not business domains.  It
never creates a logical dataset, scenario binding, semantic mapping or legacy
data source.  A caller must explicitly select an existing tenant-owned managed
file bucket, while the server derives all physical MinIO coordinates.
"""
from __future__ import annotations

import csv
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
import hashlib
from io import BytesIO, StringIO
import json
from pathlib import Path, PurePosixPath
import re
import threading
from typing import Any, Iterable, Iterator
import uuid
from zipfile import BadZipFile, ZipFile

from sqlalchemy import select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from ..config import get_settings
from ..catalog_schemas import (
    CatalogManagedUploadMetadata,
    DataAssetCreate,
    DataAssetVersionRegister,
)
from ..models import BucketFile, DataAsset, DataAssetVersion, DataSource
from . import (
    catalog_service,
    datasource_service,
    doc_parser,
    object_deletion_service,
    permission_service,
    tenant_service,
)


PROFILE_FORMAT = "catalog-file-profile/v1"
DEFAULT_ATTACHMENT_TTL_SECONDS = 24 * 60 * 60
MAX_PROFILE_COLUMNS = 500
MAX_PROFILE_ROWS = 1_000
MAX_ZIP_MEMBERS = 10_000
MAX_ZIP_EXPANDED_BYTES = 200 * 1024 * 1024
MAX_ZIP_COMPRESSION_RATIO = 250
_UPLOAD_LOCK_STRIPES = tuple(threading.RLock() for _ in range(128))

_INTEGER_RE = re.compile(r"[+-]?(?:0|[1-9][0-9]*)")
_DATE_RE = re.compile(r"\d{4}-\d{2}-\d{2}")
_DATETIME_RE = re.compile(
    r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}(?::\d{2}(?:\.\d+)?)?(?:Z|[+-]\d{2}:?\d{2})?"
)
_RELATION_NAME_CLEAN_RE = re.compile(r"[\x00-\x1f\x7f]")
_RESERVED_LABELS = {
    "catalog_purpose",
    "temporary",
    "expires_at",
    "promotion_policy",
}
_PHYSICAL_METADATA_KEYS = {
    "access_key",
    "bucket",
    "bucket_name",
    "credential",
    "credentials",
    "endpoint",
    "file_bucket_id",
    "object_key",
    "object_path",
    "password",
    "path",
    "prefix",
    "secret_key",
    "storage_path",
    "token",
}


@dataclass(frozen=True)
class FormatSpec:
    extension: str
    media_type: str
    category: str
    compatible_media_types: frozenset[str] = frozenset()


_FORMAT_SPECS: dict[str, FormatSpec] = {
    ".csv": FormatSpec(
        ".csv",
        "text/csv",
        "table",
        frozenset({"application/vnd.ms-excel", "text/plain"}),
    ),
    ".tsv": FormatSpec(
        ".tsv", "text/tab-separated-values", "table", frozenset({"text/plain"})
    ),
    ".xls": FormatSpec(".xls", "application/vnd.ms-excel", "table"),
    ".xlsx": FormatSpec(
        ".xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "table",
        frozenset({"application/zip"}),
    ),
    ".xlsm": FormatSpec(
        ".xlsm",
        "application/vnd.ms-excel.sheet.macroenabled.12",
        "table",
        frozenset({"application/zip"}),
    ),
    ".docx": FormatSpec(
        ".docx",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "document",
        frozenset({"application/zip"}),
    ),
    ".pptx": FormatSpec(
        ".pptx",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "document",
        frozenset({"application/zip"}),
    ),
    ".pdf": FormatSpec(".pdf", "application/pdf", "document"),
    ".txt": FormatSpec(".txt", "text/plain", "document"),
    ".md": FormatSpec(".md", "text/markdown", "document", frozenset({"text/plain"})),
    ".markdown": FormatSpec(
        ".markdown", "text/markdown", "document", frozenset({"text/plain"})
    ),
    ".json": FormatSpec(
        ".json", "application/json", "document", frozenset({"text/json", "text/plain"})
    ),
    ".yaml": FormatSpec(
        ".yaml", "application/yaml", "document", frozenset({"text/yaml", "text/plain"})
    ),
    ".yml": FormatSpec(
        ".yml", "application/yaml", "document", frozenset({"text/yaml", "text/plain"})
    ),
    ".xml": FormatSpec(
        ".xml", "application/xml", "document", frozenset({"text/xml", "text/plain"})
    ),
    ".log": FormatSpec(".log", "text/plain", "document"),
    ".png": FormatSpec(".png", "image/png", "document"),
    ".jpg": FormatSpec(".jpg", "image/jpeg", "document"),
    ".jpeg": FormatSpec(".jpeg", "image/jpeg", "document"),
    ".gif": FormatSpec(".gif", "image/gif", "document"),
    ".bmp": FormatSpec(".bmp", "image/bmp", "document"),
    ".tif": FormatSpec(".tif", "image/tiff", "document"),
    ".tiff": FormatSpec(".tiff", "image/tiff", "document"),
    ".webp": FormatSpec(".webp", "image/webp", "document"),
}


@dataclass(frozen=True)
class PreparedCatalogUpload:
    content: bytes
    filename: str
    content_sha256: str
    media_type: str
    profile: dict[str, Any]
    metadata: CatalogManagedUploadMetadata
    asset_key: str
    asset_name: str
    labels: dict[str, Any]
    lifecycle: dict[str, Any]
    expires_at: datetime | None

    @property
    def temporary(self) -> bool:
        return self.metadata.purpose == "invocation_attachment"


@dataclass(frozen=True)
class PreparedCatalogPathUpload:
    source_path: Path
    byte_size: int
    filename: str
    content_sha256: str
    media_type: str
    profile: dict[str, Any]
    metadata: CatalogManagedUploadMetadata
    asset_key: str
    asset_name: str
    labels: dict[str, Any]
    lifecycle: dict[str, Any]
    expires_at: datetime | None

    @property
    def temporary(self) -> bool:
        return self.metadata.purpose == "invocation_attachment"


@dataclass(frozen=True)
class ManagedCatalogUploadResult:
    asset_id: str
    version_id: str
    created: bool


@contextmanager
def _serialize_upload_identity(
    db: Session,
    *,
    tenant_id: str,
    asset_key: str,
) -> Iterator[None]:
    """Serialize one logical upload across workers before deduplication.

    PostgreSQL advisory transaction locks cover multiple API processes. The
    striped lock keeps SQLite/unit-test behavior deterministic without growing
    a process-global lock registry for attacker-controlled keys.
    """
    identity = hashlib.sha256(
        b"ontology-platform/catalog-upload-lock/v1\0"
        + tenant_id.encode("utf-8")
        + b"\0"
        + asset_key.encode("utf-8")
    ).digest()
    if db.get_bind().dialect.name == "postgresql":
        lock_key = int.from_bytes(identity[:8], byteorder="big", signed=True)
        db.execute(text("SELECT pg_advisory_xact_lock(:lock_key)"), {"lock_key": lock_key})
        yield
        return
    lock = _UPLOAD_LOCK_STRIPES[
        int.from_bytes(identity[:2], "big") % len(_UPLOAD_LOCK_STRIPES)
    ]
    with lock:
        yield


def _media_base(value: str | None) -> str:
    return str(value or "").split(";", 1)[0].strip().lower()


def _format_spec(filename: str, client_media_type: str | None) -> FormatSpec:
    extension = Path(filename).suffix.lower()
    spec = _FORMAT_SPECS.get(extension)
    if spec is None:
        raise catalog_service.CatalogError(f"不支持的目录上传文件类型: {extension or '无扩展名'}")
    supplied = _media_base(client_media_type)
    accepted = {
        _media_base(spec.media_type),
        *(_media_base(item) for item in spec.compatible_media_types),
        "application/octet-stream",
        "",
    }
    if supplied not in accepted:
        raise catalog_service.CatalogError("文件 MIME 与扩展名不一致")
    return spec


def _validate_zip_container(content: bytes, required_prefix: str) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            infos = archive.infolist()
            if not infos or len(infos) > MAX_ZIP_MEMBERS:
                raise catalog_service.CatalogError("Office 文件容器条目数量异常")
            expanded = 0
            compressed = 0
            names: set[str] = set()
            for info in infos:
                pure = PurePosixPath(info.filename.replace("\\", "/"))
                if pure.is_absolute() or ".." in pure.parts or info.flag_bits & 0x1:
                    raise catalog_service.CatalogError("Office 文件容器包含不安全条目")
                expanded += max(0, int(info.file_size))
                compressed += max(0, int(info.compress_size))
                names.add(str(pure))
            if expanded > MAX_ZIP_EXPANDED_BYTES:
                raise catalog_service.CatalogError("Office 文件展开后超过安全限制")
            if (
                expanded > 10 * 1024 * 1024
                and expanded > max(1, compressed) * MAX_ZIP_COMPRESSION_RATIO
            ):
                raise catalog_service.CatalogError("Office 文件压缩比超过安全限制")
            if not any(name.startswith(required_prefix) for name in names):
                raise catalog_service.CatalogError("文件内容与 Office 扩展名不一致")
    except BadZipFile as exc:
        raise catalog_service.CatalogError("Office 文件容器无效") from exc


def _validate_zip_container_path(path: Path, required_prefix: str) -> None:
    """Validate an Office ZIP from disk without expanding it into memory."""
    try:
        with ZipFile(path) as archive:
            infos = archive.infolist()
            if not infos or len(infos) > MAX_ZIP_MEMBERS:
                raise catalog_service.CatalogError("Office 文件容器条目数量异常")
            expanded = 0
            compressed = 0
            found_required = False
            for info in infos:
                pure = PurePosixPath(info.filename.replace("\\", "/"))
                if pure.is_absolute() or ".." in pure.parts or info.flag_bits & 0x1:
                    raise catalog_service.CatalogError("Office 文件容器包含不安全条目")
                expanded += max(0, int(info.file_size))
                compressed += max(0, int(info.compress_size))
                found_required = found_required or str(pure).startswith(required_prefix)
            expanded_limit = int(get_settings().catalog_max_office_expanded_bytes)
            if expanded > expanded_limit:
                raise catalog_service.CatalogError(
                    f"Office 文件展开后超过安全限制（{expanded_limit // (1024 * 1024)} MB）"
                )
            if (
                expanded > 10 * 1024 * 1024
                and expanded > max(1, compressed) * MAX_ZIP_COMPRESSION_RATIO
            ):
                raise catalog_service.CatalogError("Office 文件压缩比超过安全限制")
            if not found_required:
                raise catalog_service.CatalogError("文件内容与 Office 扩展名不一致")
    except BadZipFile as exc:
        raise catalog_service.CatalogError("Office 文件容器无效") from exc


def _validate_signature(content: bytes, spec: FormatSpec) -> None:
    ext = spec.extension
    if ext in {".xlsx", ".xlsm"}:
        _validate_zip_container(content, "xl/")
    elif ext == ".docx":
        _validate_zip_container(content, "word/")
    elif ext == ".pptx":
        _validate_zip_container(content, "ppt/")
    elif ext == ".xls" and not content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
        raise catalog_service.CatalogError("文件内容与 .xls 扩展名不一致")
    elif ext == ".pdf" and not content.lstrip().startswith(b"%PDF-"):
        raise catalog_service.CatalogError("文件内容与 PDF 扩展名不一致")
    elif ext == ".png" and not content.startswith(b"\x89PNG\r\n\x1a\n"):
        raise catalog_service.CatalogError("文件内容与 PNG 扩展名不一致")
    elif ext in {".jpg", ".jpeg"} and not content.startswith(b"\xff\xd8\xff"):
        raise catalog_service.CatalogError("文件内容与 JPEG 扩展名不一致")
    elif ext == ".gif" and not content.startswith((b"GIF87a", b"GIF89a")):
        raise catalog_service.CatalogError("文件内容与 GIF 扩展名不一致")
    elif ext == ".bmp" and not content.startswith(b"BM"):
        raise catalog_service.CatalogError("文件内容与 BMP 扩展名不一致")
    elif ext in {".tif", ".tiff"} and not content.startswith((b"II*\x00", b"MM\x00*")):
        raise catalog_service.CatalogError("文件内容与 TIFF 扩展名不一致")
    elif ext == ".webp" and not (
        content.startswith(b"RIFF") and len(content) >= 12 and content[8:12] == b"WEBP"
    ):
        raise catalog_service.CatalogError("文件内容与 WebP 扩展名不一致")


def _decode_text(content: bytes) -> str:
    if content.startswith((b"\xff\xfe", b"\xfe\xff")):
        try:
            return content.decode("utf-16")
        except UnicodeDecodeError as exc:
            raise catalog_service.CatalogError("UTF-16 文本文件编码无效") from exc
    if b"\x00" in content[:8192]:
        raise catalog_service.CatalogError("文本文件包含二进制内容")
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise catalog_service.CatalogError("文本文件编码不受支持")


def _logical_type(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        lowered = raw.lower()
        if lowered in {"true", "false"}:
            return "boolean"
        if _INTEGER_RE.fullmatch(raw):
            return "integer"
        try:
            Decimal(raw)
            return "number"
        except InvalidOperation:
            pass
        if _DATETIME_RE.fullmatch(raw):
            return "datetime"
        if _DATE_RE.fullmatch(raw):
            return "date"
        return "string"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, (float, Decimal)):
        return "number"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    return "string"


def _merged_type(observed: Iterable[str]) -> str:
    kinds = set(observed)
    if not kinds:
        return "unknown"
    if kinds <= {"integer", "number"}:
        return "number" if "number" in kinds else "integer"
    if kinds <= {"date", "datetime"}:
        return "datetime" if "datetime" in kinds else "date"
    if len(kinds) == 1:
        return next(iter(kinds))
    return "string"


def _table_contract(name: str, header: list[Any], rows: list[list[Any]], *, truncated: bool) -> dict[str, Any]:
    width = max([len(header), *(len(row) for row in rows)], default=0)
    if width <= 0:
        raise catalog_service.CatalogError("表格没有可识别的列")
    if width > MAX_PROFILE_COLUMNS:
        raise catalog_service.CatalogError(f"表格列数不能超过 {MAX_PROFILE_COLUMNS}")
    columns: list[dict[str, Any]] = []
    for index in range(width):
        source_header = header[index] if index < len(header) else None
        title = str(source_header).strip() if source_header is not None else ""
        if len(title) > 500:
            raise catalog_service.CatalogError("表头名称过长")
        observed: list[str] = []
        null_count = 0
        for row in rows:
            value = row[index] if index < len(row) else None
            kind = _logical_type(value)
            if kind is None:
                null_count += 1
            else:
                observed.append(kind)
        columns.append(
            {
                "ordinal": index,
                "name": title or f"column_{index + 1}",
                "logical_type": _merged_type(observed),
                "nullable": not rows or null_count > 0,
                "observed_non_null_count": len(observed),
            }
        )
    return {
        "name": str(name)[:300] or "table",
        "columns": columns,
        "sample_row_count": len(rows),
        "sample_truncated": truncated,
    }


def runtime_relation_name(
    filename: str,
    source_relation_name: str,
    relation_count: int,
) -> str:
    """Return one logical relation identity shared by modeling and runtime."""

    stem = Path(filename).stem.strip() or "table"
    source = str(source_relation_name or "").strip() or "table"
    value = stem if int(relation_count or 0) <= 1 else f"{stem}__{source}"
    value = _RELATION_NAME_CLEAN_RE.sub("", value).strip() or "table"
    if len(value) > 180:
        digest = hashlib.sha256(value.encode("utf-8")).hexdigest()[:16]
        value = f"{value[:160]}-{digest}"
    return value


def _with_runtime_relation_names(profile: dict[str, Any], filename: str) -> dict[str, Any]:
    if profile.get("category") != "table":
        return profile
    tables = [item for item in profile.get("tables") or [] if isinstance(item, dict)]
    for table in tables:
        table["relation_name"] = runtime_relation_name(
            filename,
            str(table.get("name") or ""),
            len(tables),
        )
    return profile


def _csv_profile(content: bytes, spec: FormatSpec) -> dict[str, Any]:
    text = _decode_text(content)
    sample = text[:65536]
    allowed_delimiters = "\t,;|"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=allowed_delimiters)
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if spec.extension == ".tsv" else ","
    reader = csv.reader(StringIO(text, newline=""), delimiter=delimiter)
    header: list[Any] | None = None
    rows: list[list[Any]] = []
    truncated = False
    for row in reader:
        if not any(str(value).strip() for value in row):
            continue
        if header is None:
            header = list(row)
            continue
        if len(rows) >= MAX_PROFILE_ROWS:
            truncated = True
            break
        rows.append(list(row))
    if header is None:
        raise catalog_service.CatalogError("CSV/TSV 文件为空")
    return {
        "format": PROFILE_FORMAT,
        "category": "table",
        "extension": spec.extension,
        "media_type": spec.media_type,
        "tables": [{**_table_contract("data", header, rows, truncated=truncated), "header_row_index": 0}],
    }


def _sheet_profile_rows(
    values: Iterable[Iterable[Any]],
) -> tuple[int, list[Any], list[list[Any]], bool] | None:
    """Detect a header after optional title rows, then retain a bounded sample."""
    buffered: list[tuple[int, list[Any]]] = []
    iterator = iter(values)
    for raw_index, raw in enumerate(iterator):
        row = list(raw)
        if any(_logical_type(value) is not None for value in row):
            buffered.append((raw_index, row))
        if len(buffered) >= 25:
            break
    if not buffered:
        return None
    header_position = max(
        range(len(buffered)),
        key=lambda index: sum(
            _logical_type(value) is not None for value in buffered[index][1]
        ),
    )
    header_index, header = buffered[header_position]
    rows = [row for _index, row in buffered[header_position + 1 :]]
    truncated = False
    for raw in iterator:
        row = list(raw)
        if not any(_logical_type(value) is not None for value in row):
            continue
        if len(rows) >= MAX_PROFILE_ROWS:
            truncated = True
            break
        rows.append(row)
    return header_index, header, rows[:MAX_PROFILE_ROWS], truncated


def _xlsx_profile(content: bytes, spec: FormatSpec) -> dict[str, Any]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
            keep_vba=False,
        )
    except Exception as exc:  # noqa: BLE001 - parser details are intentionally hidden.
        raise catalog_service.CatalogError("Excel 文件解析失败") from exc
    tables: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            detected = _sheet_profile_rows(sheet.iter_rows(values_only=True))
            if detected is not None:
                header_index, header, rows, truncated = detected
                tables.append(
                    {
                        **_table_contract(sheet.title, header, rows, truncated=truncated),
                        "header_row_index": header_index,
                    }
                )
    finally:
        workbook.close()
    if not tables:
        raise catalog_service.CatalogError("Excel 文件没有可识别的工作表")
    return {
        "format": PROFILE_FORMAT,
        "category": "table",
        "extension": spec.extension,
        "media_type": spec.media_type,
        "tables": tables,
    }


def _xls_profile(content: bytes, spec: FormatSpec) -> dict[str, Any]:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise catalog_service.CatalogError("服务端缺少 .xls 解析依赖") from exc
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as exc:  # noqa: BLE001
        raise catalog_service.CatalogError("Excel 文件解析失败") from exc
    tables: list[dict[str, Any]] = []

    def value_at(sheet, row: int, column: int) -> Any:
        cell = sheet.cell(row, column)
        if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
            return None
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate.xldate_as_datetime(
                    cell.value, workbook.datemode
                )
            except (ValueError, OverflowError):
                return cell.value
        return cell.value

    try:
        for sheet in workbook.sheets():
            if sheet.nrows <= 0:
                continue
            header = [value_at(sheet, 0, column) for column in range(sheet.ncols)]
            count = min(max(0, sheet.nrows - 1), MAX_PROFILE_ROWS)
            rows = [
                [value_at(sheet, row, column) for column in range(sheet.ncols)]
                for row in range(1, 1 + count)
            ]
            tables.append(
                _table_contract(
                    sheet.name,
                    header,
                    rows,
                    truncated=sheet.nrows - 1 > MAX_PROFILE_ROWS,
                )
            )
    finally:
        workbook.release_resources()
    if not tables:
        raise catalog_service.CatalogError("Excel 文件没有可识别的工作表")
    return {
        "format": PROFILE_FORMAT,
        "category": "table",
        "extension": spec.extension,
        "media_type": spec.media_type,
        "tables": tables,
    }


def _document_profile(content: bytes, filename: str, spec: FormatSpec) -> dict[str, Any]:
    if spec.extension in {".txt", ".md", ".markdown", ".json", ".yaml", ".yml", ".xml", ".log"}:
        _decode_text(content)
    parsed = doc_parser.parse_bytes(content, filename)
    succeeded = parsed.get("status") == "success"
    text = str(parsed.get("text") or "") if succeeded else ""
    return {
        "format": PROFILE_FORMAT,
        "category": "document",
        "extension": spec.extension,
        "media_type": spec.media_type,
        "parser": {
            "name": "platform-document-parser",
            "status": "parsed" if succeeded else "unavailable",
        },
        "text": {
            "character_count": len(text),
            "line_count": len(text.splitlines()) if text else 0,
        },
    }


def build_profile(content: bytes, filename: str, client_media_type: str | None = None) -> tuple[str, dict[str, Any]]:
    if not content:
        raise catalog_service.CatalogError("上传文件不能为空")
    safe_name = datasource_service.validate_bucket_filename(filename)
    spec = _format_spec(safe_name, client_media_type)
    _validate_signature(content, spec)
    if spec.extension in {".csv", ".tsv"}:
        profile = _csv_profile(content, spec)
    elif spec.extension in {".xlsx", ".xlsm"}:
        profile = _xlsx_profile(content, spec)
    elif spec.extension == ".xls":
        profile = _xls_profile(content, spec)
    else:
        profile = _document_profile(content, safe_name, spec)
    profile = _with_runtime_relation_names(profile, safe_name)
    # Reuse the catalog document guard to guarantee profiles cannot grow into
    # raw row storage or accidentally contain credential-shaped keys.
    profile = catalog_service.safe_catalog_document(
        profile, label="文件结构 profile", maximum=128_000
    )
    return spec.media_type, profile


def _text_encoding_from_sample(sample: bytes) -> str:
    if sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16"
    if b"\x00" in sample[:8192]:
        raise catalog_service.CatalogError("文本文件包含二进制内容")
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    raise catalog_service.CatalogError("文本文件编码不受支持")


def _csv_profile_path(path: Path, spec: FormatSpec) -> dict[str, Any]:
    with path.open("rb") as raw:
        sample_bytes = raw.read(65536)
    encoding = _text_encoding_from_sample(sample_bytes)
    sample = sample_bytes.decode(encoding)
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters="\t,;|").delimiter
    except csv.Error:
        delimiter = "\t" if spec.extension == ".tsv" else ","
    header: list[Any] | None = None
    rows: list[list[Any]] = []
    truncated = False
    try:
        with path.open("r", encoding=encoding, newline="") as handle:
            for row in csv.reader(handle, delimiter=delimiter):
                if not any(str(value).strip() for value in row):
                    continue
                if header is None:
                    header = list(row)
                    continue
                if len(rows) >= MAX_PROFILE_ROWS:
                    truncated = True
                    break
                rows.append(list(row))
    except UnicodeDecodeError as exc:
        raise catalog_service.CatalogError("CSV/TSV 文件编码无效") from exc
    if header is None:
        raise catalog_service.CatalogError("CSV/TSV 文件为空")
    return {
        "format": PROFILE_FORMAT,
        "category": "table",
        "extension": spec.extension,
        "media_type": spec.media_type,
        "tables": [{**_table_contract("data", header, rows, truncated=truncated), "header_row_index": 0}],
    }


def _xlsx_profile_path(path: Path, spec: FormatSpec) -> dict[str, Any]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
            keep_vba=False,
        )
    except Exception as exc:  # noqa: BLE001
        raise catalog_service.CatalogError("Excel 文件解析失败") from exc
    tables: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            detected = _sheet_profile_rows(sheet.iter_rows(values_only=True))
            if detected is not None:
                header_index, header, rows, truncated = detected
                tables.append(
                    {
                        **_table_contract(sheet.title, header, rows, truncated=truncated),
                        "header_row_index": header_index,
                    }
                )
    finally:
        workbook.close()
    if not tables:
        raise catalog_service.CatalogError("Excel 文件没有可识别的工作表")
    return {
        "format": PROFILE_FORMAT,
        "category": "table",
        "extension": spec.extension,
        "media_type": spec.media_type,
        "tables": tables,
    }


def _validate_signature_path(path: Path, spec: FormatSpec) -> None:
    if spec.extension in {".xlsx", ".xlsm"}:
        _validate_zip_container_path(path, "xl/")
        return
    if spec.extension == ".docx":
        _validate_zip_container_path(path, "word/")
        return
    if spec.extension == ".pptx":
        _validate_zip_container_path(path, "ppt/")
        return
    with path.open("rb") as handle:
        prefix = handle.read(16)
    _validate_signature(prefix, spec)


def build_profile_path(
    source_path: str | Path,
    filename: str,
    client_media_type: str | None = None,
) -> tuple[str, dict[str, Any]]:
    """Build a bounded schema/profile from a staged file in constant memory."""
    path = Path(source_path).resolve(strict=True)
    if not path.is_file() or path.stat().st_size <= 0:
        raise catalog_service.CatalogError("上传文件不能为空")
    safe_name = datasource_service.validate_bucket_filename(filename)
    spec = _format_spec(safe_name, client_media_type)
    _validate_signature_path(path, spec)
    if spec.extension in {".csv", ".tsv"}:
        profile = _csv_profile_path(path, spec)
    elif spec.extension in {".xlsx", ".xlsm"}:
        profile = _xlsx_profile_path(path, spec)
    elif spec.extension == ".xls":
        try:
            import xlrd  # type: ignore

            workbook = xlrd.open_workbook(filename=str(path), on_demand=True)
            tables = []
            try:
                for sheet in workbook.sheets():
                    if sheet.nrows <= 0:
                        continue
                    header = [sheet.cell_value(0, column) for column in range(sheet.ncols)]
                    count = min(max(0, sheet.nrows - 1), MAX_PROFILE_ROWS)
                    rows = [
                        [sheet.cell_value(row, column) for column in range(sheet.ncols)]
                        for row in range(1, 1 + count)
                    ]
                    tables.append(
                        _table_contract(
                            sheet.name,
                            header,
                            rows,
                            truncated=sheet.nrows - 1 > MAX_PROFILE_ROWS,
                        )
                    )
            finally:
                workbook.release_resources()
        except ImportError as exc:
            raise catalog_service.CatalogError("服务端缺少 .xls 解析依赖") from exc
        except Exception as exc:  # noqa: BLE001
            raise catalog_service.CatalogError("Excel 文件解析失败") from exc
        if not tables:
            raise catalog_service.CatalogError("Excel 文件没有可识别的工作表")
        profile = {
            "format": PROFILE_FORMAT,
            "category": "table",
            "extension": spec.extension,
            "media_type": spec.media_type,
            "tables": tables,
        }
    else:
        in_memory_limit = int(get_settings().catalog_in_memory_upload_bytes)
        if path.stat().st_size <= in_memory_limit:
            _media, profile = build_profile(path.read_bytes(), safe_name, client_media_type)
        else:
            profile = {
                "format": PROFILE_FORMAT,
                "category": "document",
                "extension": spec.extension,
                "media_type": spec.media_type,
                "parser": {"name": "platform-document-parser", "status": "deferred"},
                "text": {"character_count": 0, "line_count": 0},
            }
    profile = _with_runtime_relation_names(profile, safe_name)
    profile = catalog_service.safe_catalog_document(
        profile, label="文件结构 profile", maximum=128_000
    )
    return spec.media_type, profile


def profile_summary_text(profile: dict[str, Any], filename: str) -> str:
    """Render schema-only modeling context; profiles never contain business rows."""
    if not isinstance(profile, dict) or profile.get("category") != "table":
        return ""
    tables: list[dict[str, Any]] = []
    for raw_table in profile.get("tables") or []:
        if not isinstance(raw_table, dict):
            continue
        tables.append(
            {
                "name": str(
                    raw_table.get("relation_name")
                    or raw_table.get("name")
                    or "table"
                ),
                "sample_row_count": int(raw_table.get("sample_row_count") or 0),
                "sample_truncated": bool(raw_table.get("sample_truncated", False)),
                "columns": [
                    {
                        "name": str(column.get("name") or ""),
                        "logical_type": str(column.get("logical_type") or "unknown"),
                        "nullable": bool(column.get("nullable", True)),
                    }
                    for column in (raw_table.get("columns") or [])
                    if isinstance(column, dict)
                ],
            }
        )
    return (
        f"【表格结构：{datasource_service.validate_bucket_filename(filename)}】\n"
        "原始业务行保存在 MinIO，建模上下文只包含以下结构元数据：\n"
        + json.dumps({"tables": tables}, ensure_ascii=False, indent=2)
    )


def prepare_upload(
    content: bytes,
    filename: str,
    client_media_type: str | None,
    metadata: CatalogManagedUploadMetadata,
    *,
    now: datetime | None = None,
) -> PreparedCatalogUpload:
    safe_name = datasource_service.validate_bucket_filename(filename)
    media_type, profile = build_profile(content, safe_name, client_media_type)
    digest = hashlib.sha256(content).hexdigest()
    labels = catalog_service.safe_catalog_document(
        metadata.labels, label="资产标签", maximum=32_000
    )

    def reject_physical_metadata(node: Any) -> None:
        if isinstance(node, dict):
            for raw_key, value in node.items():
                key = str(raw_key).strip().lower().replace("-", "_")
                if key in _PHYSICAL_METADATA_KEYS:
                    raise catalog_service.CatalogError(
                        "资产标签不得包含对象路径、存储配置或凭据字段"
                    )
                reject_physical_metadata(value)
        elif isinstance(node, list):
            for value in node:
                reject_physical_metadata(value)

    reject_physical_metadata(labels)
    conflicting = _RESERVED_LABELS.intersection(
        str(key).strip().lower() for key in labels
    )
    if conflicting:
        raise catalog_service.CatalogError("资产标签不得覆盖平台生命周期字段")
    clock = now or datetime.now(timezone.utc)
    expires_at = None
    if metadata.purpose == "invocation_attachment":
        expires_at = clock + timedelta(
            seconds=metadata.expires_in_seconds or DEFAULT_ATTACHMENT_TTL_SECONDS
        )
    lifecycle = {
        "purpose": metadata.purpose,
        "temporary": metadata.purpose == "invocation_attachment",
        "expires_at": expires_at.isoformat() if expires_at else None,
        "promotion_policy": (
            "explicit_copy_required"
            if metadata.purpose == "invocation_attachment"
            else "not_applicable"
        ),
        "auto_promote": False,
    }
    asset_key = metadata.asset_key or f"upload.{metadata.purpose}.{digest}"
    # Validate through the same catalog key policy used by ordinary assets.
    asset_key = catalog_service.validate_catalog_key(asset_key, "资产 key")
    labels = {
        **labels,
        "catalog_purpose": metadata.purpose,
        "temporary": lifecycle["temporary"],
        "promotion_policy": lifecycle["promotion_policy"],
    }
    if expires_at:
        labels["expires_at"] = expires_at.isoformat()
    return PreparedCatalogUpload(
        content=content,
        filename=safe_name,
        content_sha256=digest,
        media_type=media_type,
        profile=profile,
        metadata=metadata,
        asset_key=asset_key,
        asset_name=(metadata.name or safe_name).strip(),
        labels=labels,
        lifecycle=lifecycle,
        expires_at=expires_at,
    )


def prepare_upload_path(
    source_path: str | Path,
    filename: str,
    client_media_type: str | None,
    metadata: CatalogManagedUploadMetadata,
    *,
    content_sha256: str,
    byte_size: int,
    now: datetime | None = None,
) -> PreparedCatalogPathUpload:
    path = Path(source_path).resolve(strict=True)
    size = int(byte_size)
    if not path.is_file() or size <= 0 or path.stat().st_size != size:
        raise catalog_service.CatalogError("上传暂存文件大小无效")
    digest = str(content_sha256 or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", digest):
        raise catalog_service.CatalogError("上传文件缺少有效内容哈希")
    safe_name = datasource_service.validate_bucket_filename(filename)
    media_type, profile = build_profile_path(path, safe_name, client_media_type)
    labels = catalog_service.safe_catalog_document(
        metadata.labels, label="资产标签", maximum=32_000
    )

    def reject_physical_metadata(node: Any) -> None:
        if isinstance(node, dict):
            for raw_key, value in node.items():
                key = str(raw_key).strip().lower().replace("-", "_")
                if key in _PHYSICAL_METADATA_KEYS:
                    raise catalog_service.CatalogError(
                        "资产标签不得包含对象路径、存储配置或凭据字段"
                    )
                reject_physical_metadata(value)
        elif isinstance(node, list):
            for value in node:
                reject_physical_metadata(value)

    reject_physical_metadata(labels)
    if _RESERVED_LABELS.intersection(str(key).strip().lower() for key in labels):
        raise catalog_service.CatalogError("资产标签不得覆盖平台生命周期字段")
    clock = now or datetime.now(timezone.utc)
    expires_at = None
    if metadata.purpose == "invocation_attachment":
        expires_at = clock + timedelta(
            seconds=metadata.expires_in_seconds or DEFAULT_ATTACHMENT_TTL_SECONDS
        )
    temporary = metadata.purpose == "invocation_attachment"
    lifecycle = {
        "purpose": metadata.purpose,
        "temporary": temporary,
        "expires_at": expires_at.isoformat() if expires_at else None,
        "promotion_policy": "explicit_copy_required" if temporary else "not_applicable",
        "auto_promote": False,
    }
    asset_key = catalog_service.validate_catalog_key(
        metadata.asset_key or f"upload.{metadata.purpose}.{digest}", "资产 key"
    )
    labels = {
        **labels,
        "catalog_purpose": metadata.purpose,
        "temporary": temporary,
        "promotion_policy": lifecycle["promotion_policy"],
    }
    if expires_at:
        labels["expires_at"] = expires_at.isoformat()
    return PreparedCatalogPathUpload(
        source_path=path,
        byte_size=size,
        filename=safe_name,
        content_sha256=digest,
        media_type=media_type,
        profile=profile,
        metadata=metadata,
        asset_key=asset_key,
        asset_name=(metadata.name or safe_name).strip(),
        labels=labels,
        lifecycle=lifecycle,
        expires_at=expires_at,
    )


def require_managed_file_bucket(db: Session, file_bucket_id: str) -> DataSource:
    permission_service.require_tenant_permission(db, "write")
    source = db.scalar(
        select(DataSource)
        .where(
            DataSource.id == file_bucket_id,
            DataSource.tenant_id == tenant_service.current_tenant_id(db),
        )
        .execution_options(populate_existing=True)
        .with_for_update()
    )
    if source is None:
        raise catalog_service.CatalogError("受管文件桶不存在")
    if source.scenario_id:
        scenario = tenant_service.require_scenario(db, source.scenario_id, writable=True)
        permission_service.require_scenario_permission(
            db,
            scenario,
            "write",
            message="没有所选文件桶所属业务场景的权限",
        )
    if not datasource_service.is_managed_minio_source(source):
        raise catalog_service.CatalogError("只能选择服务端受管 MinIO 文件桶")
    try:
        datasource_service.managed_minio_location(source)
    except Exception as exc:  # noqa: BLE001 - do not reveal configuration details.
        raise catalog_service.CatalogError("所选文件桶不符合平台受管存储策略") from exc
    return source


def require_external_upload_bucket(db: Session) -> DataSource:
    """Resolve the tenant-owned upload bucket without exposing physical identity."""
    permission_service.require_tenant_permission(db, "write")
    tenant_id = tenant_service.current_tenant_id(db)
    source_id = hashlib.sha256(
        b"ontology-platform/external-upload-source/v1\0"
        + tenant_id.encode("utf-8")
    ).hexdigest()[:32]
    source = db.scalar(
        select(DataSource)
        .where(
            DataSource.id == source_id,
            DataSource.tenant_id == tenant_id,
        )
        .execution_options(populate_existing=True)
        .with_for_update()
    )
    if source is None:
        try:
            with db.begin_nested():
                source = DataSource(
                    id=source_id,
                    tenant_id=tenant_id,
                    scenario_id=None,
                    resource_scope="agent_runtime",
                    owner_agent_id=None,
                    name="外部调用临时附件",
                    type="file_bucket",
                    config=datasource_service.normalize_file_bucket_config(),
                    status="ok",
                )
                db.add(source)
                db.flush()
        except IntegrityError:
            source = db.scalar(
                select(DataSource)
                .where(
                    DataSource.id == source_id,
                    DataSource.tenant_id == tenant_id,
                )
                .execution_options(populate_existing=True)
                .with_for_update()
            )
    if source is None or source.scenario_id is not None:
        raise catalog_service.CatalogError("外部附件托管存储不可用")
    # Databases created before the resource-scope migration used the modeling
    # default. This deterministic internal bucket is safe to reclassify and is
    # never exposed in the modeling-material page.
    source.resource_scope = "agent_runtime"
    source.owner_agent_id = None
    if not datasource_service.is_managed_minio_source(source):
        raise catalog_service.CatalogError("外部附件托管存储不可用")
    try:
        datasource_service.managed_minio_location(source)
        datasource_service.ensure_file_bucket_storage(source)
    except Exception as exc:  # noqa: BLE001 - physical configuration stays private.
        raise catalog_service.CatalogError("外部附件托管存储不可用") from exc
    return source


def find_or_create_asset(
    db: Session, prepared: PreparedCatalogUpload | PreparedCatalogPathUpload
) -> tuple[DataAsset, DataAssetVersion | None, bool, bool]:
    """Resolve content-hash idempotency before any object is uploaded."""
    reactivated = False
    existing = db.scalar(
        select(DataAsset).where(
            DataAsset.tenant_id == tenant_service.current_tenant_id(db),
            DataAsset.key == prepared.asset_key,
        )
    )
    created = existing is None
    if existing is None:
        existing = catalog_service.create_asset(
            db,
            DataAssetCreate(
                key=prepared.asset_key,
                name=prepared.asset_name,
                description=prepared.metadata.description,
                kind="file",
                media_type=prepared.media_type,
                labels=prepared.labels,
            ),
        )
    else:
        if existing.lifecycle_status != "active":
            if prepared.metadata.purpose != "validation_asset":
                raise catalog_service.CatalogError("目标数据资产已退役")
            existing.lifecycle_status = "active"
            existing.retired_at = None
            existing.labels = prepared.labels
            reactivated = True
        existing_purpose = str((existing.labels or {}).get("catalog_purpose") or "")
        if existing_purpose and existing_purpose != prepared.metadata.purpose:
            raise catalog_service.CatalogError("临时附件与长期资产不能原地互相晋级")
        if not existing_purpose:
            raise catalog_service.CatalogError("既有资产未声明上传生命周期，不能由上传接口覆盖")
        if existing.kind != "file" or existing.media_type != prepared.media_type:
            raise catalog_service.CatalogError("既有资产的文件类型与本次上传不一致")
    duplicate = db.scalar(
        select(DataAssetVersion)
        .where(
            DataAssetVersion.asset_id == existing.id,
            DataAssetVersion.content_sha256 == prepared.content_sha256,
            DataAssetVersion.status == "ready",
            DataAssetVersion.bucket_file_id.is_not(None),
        )
        .order_by(
            DataAssetVersion.version_number.desc(),
            DataAssetVersion.id.desc(),
        )
        .limit(1)
    )
    replace_expired = reactivated
    if duplicate is not None and prepared.metadata.purpose == "invocation_attachment":
        lifecycle, duplicate_expires_at = lifecycle_from_version(duplicate)
        clock = datetime.now(timezone.utc)
        duplicate_is_current = (
            lifecycle.get("purpose") == "invocation_attachment"
            and duplicate_expires_at is not None
            and (
                duplicate_expires_at
                if duplicate_expires_at.tzinfo
                else duplicate_expires_at.replace(tzinfo=timezone.utc)
            )
            > clock
            and duplicate.status == "ready"
        )
        if not duplicate_is_current:
            duplicate = None
            replace_expired = True
    if (
        not created
        and duplicate is None
        and prepared.metadata.purpose == "invocation_attachment"
        and not replace_expired
    ):
        raise catalog_service.CatalogError("临时调用附件不可追加版本或原地晋级")
    if replace_expired:
        existing.labels = {
            **dict(existing.labels or {}),
            "expires_at": prepared.lifecycle.get("expires_at"),
        }
    return existing, duplicate, created, replace_expired


def register_prepared_version(
    db: Session,
    asset: DataAsset,
    bucket_file_id: str,
    prepared: PreparedCatalogUpload | PreparedCatalogPathUpload,
    *,
    allow_duplicate_content: bool = False,
) -> DataAssetVersion:
    return catalog_service.register_asset_version(
        db,
        asset,
        DataAssetVersionRegister(
            bucket_file_id=bucket_file_id,
            provenance_kind="upload",
            version_document={
                "format": "catalog-managed-upload/v1",
                "profile": prepared.profile,
                "lifecycle": prepared.lifecycle,
            },
        ),
        allow_duplicate_content=allow_duplicate_content,
    )


def persist_managed_upload(
    db: Session,
    source: DataSource,
    content: bytes,
    filename: str,
    client_media_type: str | None,
    metadata: CatalogManagedUploadMetadata,
) -> ManagedCatalogUploadResult:
    """Atomically register one logical asset version and its exact MinIO object."""
    upload_claim = None
    bucket_file: BucketFile | None = None

    def schedule_abandoned_upload() -> None:
        if upload_claim is not None and bucket_file is not None:
            object_deletion_service.schedule_abandoned_upload_best_effort(
                upload_claim, bucket_file
            )

    try:
        prepared = prepare_upload(
            content,
            filename,
            client_media_type,
            metadata,
        )
        tenant_id = tenant_service.current_tenant_id(db)
        with _serialize_upload_identity(
            db,
            tenant_id=tenant_id,
            asset_key=prepared.asset_key,
        ):
            asset, duplicate, _asset_created, replace_expired = find_or_create_asset(
                db, prepared
            )
            if duplicate is not None:
                result = ManagedCatalogUploadResult(
                    asset_id=asset.id,
                    version_id=duplicate.id,
                    created=False,
                )
                db.rollback()
                return result

            file_id = uuid.uuid4().hex
            upload_claim = object_deletion_service.prepare_bucket_file_upload(
                source,
                file_id,
                prepared.filename,
            )
            with object_deletion_service.heartbeat_upload_intent(
                upload_claim
            ) as upload_heartbeat:
                object_deletion_service.begin_upload_put(upload_claim)
                bucket_file = datasource_service.save_bucket_file(
                    source,
                    prepared.filename,
                    prepared.content,
                    mime=prepared.media_type,
                    stable_file_id=file_id,
                    upload_object_key=upload_claim.object_key,
                )
                object_deletion_service.assert_upload_active(
                    upload_heartbeat, upload_claim, bucket_file
                )
            bucket_file.status = "parsed"
            bucket_file.error = ""
            bucket_file.parsed_text = ""
            db.add(bucket_file)
            object_deletion_service.retain_bucket_file_upload(
                db, upload_claim, bucket_file, source
            )
            db.flush()
            version = register_prepared_version(
                db,
                asset,
                bucket_file.id,
                prepared,
                allow_duplicate_content=replace_expired,
            )
            if version.bucket_file_id != bucket_file.id:
                # A writer outside this service may have won after our initial
                # lookup. Roll back its metadata and remove only our exact
                # generation-scoped object through the durable cleanup outbox.
                result = ManagedCatalogUploadResult(
                    asset_id=asset.id,
                    version_id=version.id,
                    created=False,
                )
                db.rollback()
                schedule_abandoned_upload()
                return result
            result = ManagedCatalogUploadResult(
                asset_id=asset.id,
                version_id=version.id,
                created=True,
            )
            db.commit()
            return result
    except Exception:
        db.rollback()
        schedule_abandoned_upload()
        raise


def persist_managed_upload_path(
    db: Session,
    source: DataSource,
    source_path: str | Path,
    filename: str,
    client_media_type: str | None,
    metadata: CatalogManagedUploadMetadata,
    *,
    content_sha256: str,
    byte_size: int,
) -> ManagedCatalogUploadResult:
    """Persist a large catalog upload with bounded memory and the same lifecycle fence."""
    upload_claim = None
    bucket_file: BucketFile | None = None

    def schedule_abandoned_upload() -> None:
        if upload_claim is not None and bucket_file is not None:
            object_deletion_service.schedule_abandoned_upload_best_effort(
                upload_claim, bucket_file
            )

    try:
        prepared = prepare_upload_path(
            source_path,
            filename,
            client_media_type,
            metadata,
            content_sha256=content_sha256,
            byte_size=byte_size,
        )
        tenant_id = tenant_service.current_tenant_id(db)
        with _serialize_upload_identity(
            db,
            tenant_id=tenant_id,
            asset_key=prepared.asset_key,
        ):
            asset, duplicate, _asset_created, replace_expired = find_or_create_asset(
                db, prepared
            )
            if duplicate is not None:
                result = ManagedCatalogUploadResult(
                    asset_id=asset.id,
                    version_id=duplicate.id,
                    created=False,
                )
                db.rollback()
                return result

            file_id = uuid.uuid4().hex
            upload_claim = object_deletion_service.prepare_bucket_file_upload(
                source, file_id, prepared.filename
            )
            with object_deletion_service.heartbeat_upload_intent(
                upload_claim
            ) as upload_heartbeat:
                object_deletion_service.begin_upload_put(upload_claim)
                bucket_file = datasource_service.save_bucket_file_path(
                    source,
                    prepared.filename,
                    prepared.source_path,
                    mime=prepared.media_type,
                    stable_file_id=file_id,
                    upload_object_key=upload_claim.object_key,
                    content_sha256=prepared.content_sha256,
                )
                object_deletion_service.assert_upload_active(
                    upload_heartbeat, upload_claim, bucket_file
                )
            bucket_file.status = "parsed"
            bucket_file.error = ""
            bucket_file.parsed_text = ""
            db.add(bucket_file)
            object_deletion_service.retain_bucket_file_upload(
                db, upload_claim, bucket_file, source
            )
            db.flush()
            version = register_prepared_version(
                db,
                asset,
                bucket_file.id,
                prepared,
                allow_duplicate_content=replace_expired,
            )
            if version.bucket_file_id != bucket_file.id:
                result = ManagedCatalogUploadResult(
                    asset_id=asset.id,
                    version_id=version.id,
                    created=False,
                )
                db.rollback()
                schedule_abandoned_upload()
                return result
            result = ManagedCatalogUploadResult(
                asset_id=asset.id,
                version_id=version.id,
                created=True,
            )
            db.commit()
            return result
    except Exception:
        db.rollback()
        schedule_abandoned_upload()
        raise


def lifecycle_from_version(version: DataAssetVersion) -> tuple[dict[str, Any], datetime | None]:
    lifecycle = dict((version.version_document or {}).get("lifecycle") or {})
    expires_at = None
    raw = lifecycle.get("expires_at")
    if isinstance(raw, str) and raw:
        try:
            expires_at = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            expires_at = None
    return lifecycle, expires_at


def managed_upload_document(
    asset: DataAsset,
    version: DataAssetVersion,
    *,
    fallback_purpose: str,
    created: bool,
) -> dict[str, Any]:
    """Project an upload result without exposing its physical object identity."""
    document = dict(version.version_document or {})
    lifecycle, expires_at = lifecycle_from_version(version)
    purpose = str(lifecycle.get("purpose") or fallback_purpose)
    return {
        "purpose": purpose,
        "temporary": bool(
            lifecycle.get("temporary", purpose == "invocation_attachment")
        ),
        "expires_at": expires_at,
        "created": created,
        "asset": {
            "id": asset.id,
            "key": asset.key,
            "name": asset.name,
            "kind": asset.kind,
            "media_type": asset.media_type or "",
            "lifecycle_status": asset.lifecycle_status,
        },
        "version": {
            "id": version.id,
            "asset_id": version.asset_id,
            "version_number": version.version_number,
            "provenance_kind": version.provenance_kind,
            "status": version.status,
            "content_sha256": version.content_sha256,
            "byte_size": version.byte_size,
            "profile": dict(document.get("profile") or {}),
            "lifecycle": lifecycle,
            "created_at": version.created_at,
        },
    }
